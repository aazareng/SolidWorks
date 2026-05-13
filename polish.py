#!/usr/bin/env python3
"""
polish.py — Generate a single-page HTML manual from an Excel workbook.

Two modes:

  --htm  (primary / works now)
    Post-processes Excel's "Save As HTML" export.  Reads sheetXXX.htm files,
    fixes image sizing, removes sheet-header rows, applies navigation and
    B/W/Red styling.  var_PLC field on SUMMARY auto-selects the VFD sheet.

    python polish.py --htm "manual.htm" --out output_dir

  --xlsx  (future — requires img_* named ranges in the workbook)
    Exports print areas via Excel COM automation.
    img_* named ranges → PNG (captures overlaid shapes/callouts)
    hdr_* named ranges → skipped rows
    Everything else    → clean HTML table

    python polish.py --xlsx "manual.xlsx" --out output_dir

Requirements:
  python -m pip install pywin32 openpyxl beautifulsoup4
"""

import argparse
import datetime
import re
import shutil
import sys
from pathlib import Path
from urllib.parse import unquote, quote as url_quote

try:
    import win32com.client
    import pythoncom
except ImportError:
    sys.exit("pywin32 not found.  Run:  python -m pip install pywin32")

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    sys.exit("openpyxl not found.  Run:  python -m pip install openpyxl")


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

MODULE_SHEETS = {
    "INFEED CONVEYOR", "STAGING CLAMP", "FILM CRADLE", "FILM DELIVERY",
    "TRANSFER CONVEYOR", "WRAP ARM", "TUNNEL CONVEYOR", "HEAT TUNNEL",
}
SETTINGS_SHEETS = {"HMI", "PHYSICAL", "SCHNEIDER", "AB", "POWERFLEX", "VFD"}

VFD_MAP = {
    "SCHNEIDER": "SCHNEIDER", "AB": "AB",
    "ALLEN-BRADLEY": "AB", "POWERFLEX": "POWERFLEX",
}

# Rows whose first non-empty cell matches these are sheet-header boilerplate — skip
HEADER_KEYWORDS = {
    "TECHNICAL MANUAL", "ENGINEERING MACHINE SUMMARY",
    "SERIAL NUMBER", "SERIAL NUMBER:", "TYPE", "TYPE:",
    "CUSTOMER", "CUSTOMER:", "DELIVERY DATE", "DELIVERY DATE:",
}

# BOM table column header keywords — triggers add-to-quote UI
BOM_COL_KEYWORDS = {
    "PART", "PART #", "PART NUMBER", "P/N", "PN", "ITEM #", "ITEM NO",
    "POLYPACK PN", "POLYPACK P/N", "NUMBER",
}

# Sub-section dividers within a module sheet → become sub-tabs
SUB_SECTION_KEYWORDS = {
    "OVERVIEW", "CHANGEOVER", "BOM", "MAINTENANCE",
    "ASSEMBLY", "SPARE PARTS", "LUBRICATION", "ELECTRICAL", "PNEUMATIC",
}

# Named Excel tables treated as BOM tables regardless of column headers
BOM_TABLE_PREFIX = "tbl_BOM_"
BOM_TABLE_NAMES  = {"tbl_EEBOM"}

# PM sheet "IMG" column values → inline icons (Unicode emoji, styled)
PM_IMG_ICONS = {
    "CLEAN":     "<span class='pm-icon' title='Clean'>🧹</span>",
    "SWEEP":     "<span class='pm-icon' title='Clean'>🧹</span>",
    "LUBRICATE": "<span class='pm-icon' title='Lubricate'>🛢</span>",
    "LUBE":      "<span class='pm-icon' title='Lubricate'>🛢</span>",
    "OIL":       "<span class='pm-icon' title='Lubricate'>🛢</span>",
    "GREASE":    "<span class='pm-icon' title='Grease'>🛢</span>",
    "INSPECT":   "<span class='pm-icon' title='Inspect'>🔍</span>",
    "CHECK":     "<span class='pm-icon' title='Inspect'>🔍</span>",
    "REPLACE":   "<span class='pm-icon' title='Replace'>🔧</span>",
    "TIGHTEN":   "<span class='pm-icon' title='Tighten'>🔩</span>",
}

LABELS = {
    "SUMMARY":           "Machine Summary",
    "FTD":               "Film Threading Diagram",
    "INFEED CONVEYOR":   "Infeed Conveyor",
    "STAGING CLAMP":     "Staging Clamp",
    "FILM CRADLE":       "Film Cradle",
    "FILM DELIVERY":     "Film Delivery",
    "TRANSFER CONVEYOR": "Transfer Conveyor",
    "WRAP ARM":          "Wrap Arm",
    "TUNNEL CONVEYOR":   "Tunnel Conveyor",
    "HEAT TUNNEL":       "Heat Tunnel",
    "ELE":               "Electrical",
    "FAULTS":            "Fault Codes",
    "PM":                "Preventive Maintenance",
    "TS":                "Troubleshooting",
    "HMI":               "HMI Settings",
    "SCREENS":           "HMI Screens",
    "PHYSICAL":          "Physical Setpoints",
    "LUBRICATION":       "Lubrication",
    "SCHNEIDER":         "VFD Settings (Schneider)",
    "AB":                "VFD Settings (Allen-Bradley)",
    "POWERFLEX":         "VFD Settings (PowerFlex)",
}

# Navigation groups — ordered list of (group_key, display_label)
GROUP_ORDER = [
    ("OVERVIEW",     "Machine Overview"),
    ("ELECTRICAL",   "Electrical"),
    ("SOFTWARE",     "Software"),
    ("MAINTENANCE",  "Maintenance"),
    ("SETTINGS",     "Settings"),
]

# Default group for sheets that don't have GROUP= in their A1 note.
# Sheets with type=="module" fall back to OVERVIEW automatically in main().
_DEFAULT_GROUP: dict[str, str] = {
    "SUMMARY":     "OVERVIEW",
    "FTD":         "OVERVIEW",
    "ELE":         "ELECTRICAL",
    "ELE-BOM":     "ELECTRICAL",
    "SCREENS":     "SOFTWARE",
    "FAULTS":      "SOFTWARE",
    "SCHNEIDER":   "SOFTWARE",
    "AB":          "SOFTWARE",
    "POWERFLEX":   "SOFTWARE",
    "HMI":         "SOFTWARE",
    "PHYSICAL":    "SETTINGS",
    "SETTINGS":    "SETTINGS",
    "PM":          "MAINTENANCE",
    "TS":          "MAINTENANCE",
    "LUBRICATION": "MAINTENANCE",
}

# VFD sheets are mutually exclusive — only the one matching var_PLC is included.
VFD_SHEETS = {"SCHNEIDER", "AB", "POWERFLEX"}


# ---------------------------------------------------------------------------
# Range utilities
# ---------------------------------------------------------------------------

def parse_range_ref(ref: str, max_row: int = 10000) -> tuple[int, int, int, int]:
    """
    Parse a range reference like '$A$1:$C$22', 'A1:C22', 'A22', or '$A:$G'.
    Returns (row1, row2, col1, col2) as 1-based integers.
    Column-only references like '$A:$G' use rows 1..max_row.
    """
    ref = ref.replace("$", "").strip()
    # Strip sheet name prefix like 'SUMMARY'!A1:C22
    if "!" in ref:
        ref = ref.split("!")[-1]
    if ":" in ref:
        start, end = ref.split(":")
    else:
        start = end = ref

    def col_letters_to_num(col_str: str) -> int:
        col = 0
        for ch in col_str.upper():
            col = col * 26 + (ord(ch) - ord("A") + 1)
        return col

    def cell_to_rc(cell: str) -> tuple[int, int]:
        # Full cell ref: A1, BC22, etc.
        m = re.match(r"([A-Za-z]+)(\d+)$", cell)
        if m:
            return int(m.group(2)), col_letters_to_num(m.group(1))
        # Column-only ref: A, BC, etc.
        m = re.match(r"([A-Za-z]+)$", cell)
        if m:
            return None, col_letters_to_num(m.group(1))
        raise ValueError(f"Cannot parse cell reference: {cell!r}")

    r1, c1 = cell_to_rc(start)
    r2, c2 = cell_to_rc(end)
    if r1 is None:
        r1 = 1
    if r2 is None:
        r2 = max_row
    return min(r1, r2), max(r1, r2), min(c1, c2), max(c1, c2)


def rc_to_ref(r: int, c: int) -> str:
    return f"{get_column_letter(c)}{r}"


# ---------------------------------------------------------------------------
# openpyxl helpers
# ---------------------------------------------------------------------------

def get_print_area(ws_opxl) -> str | None:
    """Return the print area string (e.g. 'A1:C22') or None."""
    pa = ws_opxl.print_area
    if not pa:
        return None
    # openpyxl may return a list or a string
    if isinstance(pa, (list, tuple)):
        pa = pa[0] if pa else None
    if pa:
        return pa.replace("$", "")
    return None


def get_named_ranges(wb_opxl, sheet_name: str, prefix: str) -> list[dict]:
    """
    Return all named ranges on sheet_name whose names start with prefix.
    Each entry: {name, ref (no sheet prefix), r1, r2, c1, c2}
    Sorted by r1 (top to bottom).
    """
    results = []
    sheet_upper = sheet_name.upper()
    prefix_lower = prefix.lower()

    for dn_name in wb_opxl.defined_names:
        if not dn_name.lower().startswith(prefix_lower):
            continue
        dn = wb_opxl.defined_names[dn_name]
        try:
            destinations = list(dn.destinations)
        except Exception:
            continue
        for dest_sheet, coord in destinations:
            if dest_sheet.upper() == sheet_upper:
                try:
                    r1, r2, c1, c2 = parse_range_ref(coord)
                    results.append({
                        "name": dn_name,
                        "ref":  coord.replace("$", ""),
                        "r1": r1, "r2": r2, "c1": c1, "c2": c2,
                    })
                except Exception as e:
                    print(f"    Warning: could not parse range {dn_name}: {e}")

    return sorted(results, key=lambda x: x["r1"])


EXCEL_ERRORS = {"#VALUE!", "#N/A", "#REF!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"}

def format_cell_value(value) -> str:
    """Convert a cell value to a display string. Suppresses Excel error strings."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.time):
        # Gear ratios stored as time e.g. 20:01 → "20:1"
        return f"{value.hour}:{value.minute}"
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    s = str(value)
    if s in EXCEL_ERRORS:
        return ""   # suppress formula errors (e.g. XLOOKUP returning image shows #VALUE!)
    return s


# ---------------------------------------------------------------------------
# openpyxl-only helpers (no COM required — VBA handles image export)
# ---------------------------------------------------------------------------

def _vba_clean_name(s: str) -> str:
    """Mirrors VBA CleanName: spaces and hyphens → underscores."""
    return s.replace(" ", "_").replace("-", "_")


def _read_sheet_meta(ws) -> dict:
    """
    Read sheet metadata from the A1 cell note (VBA: ws.Range("A1").NoteText).
    Format:  KEY=VALUE;KEY=VALUE
    Example: INCLUDE=TRUE;TYPE=MODULE;SECTIONS=OVERVIEW,CHANGEOVER,BOM
    Returns dict with uppercase keys.
    Defaults: INCLUDE=TRUE, TYPE=REGULAR, SECTIONS=OVERVIEW.
    """
    meta = {"INCLUDE": "TRUE", "TYPE": "REGULAR", "SECTIONS": "OVERVIEW"}
    try:
        comment = ws["A1"].comment
        if comment:
            txt = str(comment.text or "")
            for part in txt.split(";"):
                part = part.strip()
                if "=" in part:
                    k, v = part.split("=", 1)
                    meta[k.strip().upper()] = v.strip().upper()
    except Exception:
        pass
    return meta


def _cell_img_filename(sheet_name: str, r: int, c: int,
                       r2: int = None, c2: int = None) -> str:
    """
    Build the PNG filename VBA would produce for a cell or merged range.
    VBA: CleanName(ws.Name) & "_" & Replace(cap.Address(False,False), ":", "_") & ".png"
    Single cell A22  → SUMMARY_A22.png
    Merge A7:C7      → FTD_A7_C7.png
    """
    clean = _vba_clean_name(sheet_name)
    col1  = get_column_letter(c)
    if r2 is not None and c2 is not None and (r2 != r or c2 != c):
        addr = f"{col1}{r}_{get_column_letter(c2)}{r2}"
    else:
        addr = f"{col1}{r}"
    return f"{clean}_{addr}.png"


def _is_light_color(hex_color: str) -> bool:
    """True if color is near-white (luminance > 200/255) — skip on light-theme pages."""
    h = hex_color.lstrip("#")
    if len(h) != 6:
        return False
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return (r * 299 + g * 587 + b * 114) / 1000 > 200


def _build_assets_index(assets_dir: Path, sheet_name: str) -> dict:
    """
    Scan assets_dir for PNGs exported by VBA for sheet_name.
    Returns {(first_row, first_col): filename} so image lookup is robust to
    merge-span mismatches between VBA and openpyxl.
    """
    if not assets_dir or not assets_dir.exists():
        return {}
    clean   = _vba_clean_name(sheet_name)
    prefix  = clean + "_"
    index   = {}
    pat     = re.compile(r'^([A-Z]+)(\d+)', re.IGNORECASE)
    for f in assets_dir.iterdir():
        if f.suffix.lower() != ".png":
            continue
        name = f.name
        if not name.startswith(prefix):
            continue
        addr_part = name[len(prefix):-4]          # e.g. "A7_C7" or "A22"
        m = pat.match(addr_part)
        if m:
            try:
                col = column_index_from_string(m.group(1).upper())
                row = int(m.group(2))
                index[(row, col)] = name
            except Exception:
                pass
    return index


def load_module_images(assets_dir: Path, sheet_name: str) -> dict:
    """
    Discover section images by iterating actual subdirectories.
    Returns {section_title: img_html}  e.g. {"Overview": "<div>...", "Changeover": ...}

    Tries both the raw sheet name and the VBA-cleaned name (spaces/hyphens→_)
    so the lookup is robust regardless of which VBA CleanName variant was used.
    """
    if not assets_dir or not assets_dir.exists():
        return {}

    # Find the sheet folder — try raw name first, then cleaned
    sheet_dir = None
    for candidate in [sheet_name, _vba_clean_name(sheet_name)]:
        d = assets_dir / candidate
        if d.exists() and d.is_dir():
            sheet_dir = d
            url_sheet = candidate   # use the folder name that actually exists
            break
    if sheet_dir is None:
        return {}

    sections = {}
    for sec in sorted(sheet_dir.iterdir()):
        if not sec.is_dir():
            continue
        imgs = sorted(f.name for f in sec.iterdir() if f.suffix.lower() == ".png")
        if not imgs:
            continue
        html = "".join(
            f'<div class="section-img">'
            f'<img src="assets/{url_quote(url_sheet)}/{url_quote(sec.name)}/{url_quote(f)}"'
            f' style="max-width:100%;height:auto;display:block;margin:0 auto 1em;">'
            f'</div>'
            for f in imgs
        )
        sections[sec.name.title()] = html   # "OVERVIEW" → "Overview"

    return sections


_EXCEL_ERRORS = {"#NAME?", "#VALUE?", "#REF?", "#DIV/0!", "#N/A", "#NULL!", "#NUM!"}
_SKIP_SHEETS   = {"ADMIN", "DASHBOARD", "TEMPLATE", "CHANGELOG", "INDEX", "COVER"}

def _sheet_type_from_name(name: str) -> str:
    nu = name.upper()
    if nu == "SCREENS":       return "screens"
    if nu in MODULE_SHEETS:   return "module"
    if nu in SETTINGS_SHEETS: return "settings"
    return ""

def read_print_table_opxl(wb_opxl) -> list[dict]:
    """
    Return [{name, type}] for sheets to include in the manual, in order.

    Priority:
      1. ADMIN!tbl_PRINT — NAME column (if values are not Excel errors)
         Fallback within tbl_PRINT: Sheet column (1-based index → wb.worksheets)
      2. If tbl_PRINT missing or yields nothing valid: iterate all worksheets,
         skip system sheets and any whose A1 note has INCLUDE=FALSE.
    """
    try:
        ws_admin = wb_opxl["ADMIN"]
        for tname in ws_admin.tables:
            if tname != "tbl_PRINT":
                continue
            tobj = ws_admin.tables[tname]
            r1, r2, c1, c2 = parse_range_ref(tobj.ref)
            headers = [str(ws_admin.cell(r1, c).value or "").strip().upper()
                       for c in range(c1, c2 + 1)]
            name_col  = c1 + (headers.index("NAME")  if "NAME"  in headers else 0)
            sheet_col = (c1 + headers.index("SHEET")) if "SHEET" in headers else None
            type_col  = (c1 + headers.index("TYPE"))  if "TYPE"  in headers else None
            items = []
            for r in range(r1 + 1, r2 + 1):
                name = ""
                # Try NAME column first
                v = ws_admin.cell(r, name_col).value
                raw = str(v).strip() if v is not None else ""
                if raw and raw not in _EXCEL_ERRORS:
                    name = raw
                # Fallback: numeric Sheet column → wb.worksheets[idx-1].title
                if not name and sheet_col:
                    sv = ws_admin.cell(r, sheet_col).value
                    if sv:
                        try:
                            idx = int(sv) - 1
                            if 0 <= idx < len(wb_opxl.worksheets):
                                name = wb_opxl.worksheets[idx].title
                        except (ValueError, TypeError):
                            pass
                if not name:
                    continue
                stype = ""
                if type_col:
                    stype = str(ws_admin.cell(r, type_col).value or "").strip().lower()
                if not stype:
                    stype = _sheet_type_from_name(name)
                items.append({"name": name, "type": stype})
            if items:
                return items
    except (KeyError, Exception):
        pass

    # No tbl_PRINT or it yielded nothing — discover via A1 note INCLUDE metadata
    items = []
    for ws in wb_opxl.worksheets:
        if ws.title.upper() in _SKIP_SHEETS:
            continue
        meta = _read_sheet_meta(ws)
        if meta.get("INCLUDE", "TRUE") == "FALSE":
            continue
        stype = meta.get("TYPE", "").lower()
        if stype in ("regular", ""):
            stype = _sheet_type_from_name(ws.title)
        items.append({"name": ws.title, "type": stype})
    return items


def _resolve_named_range(wb_opxl, name: str) -> str:
    """Return the (cached) value of a workbook-scoped or sheet-scoped named range, or ''."""
    try:
        dn = wb_opxl.defined_names.get(name)
    except Exception:
        dn = None
    if dn is None:
        for ws in wb_opxl.worksheets:
            try:
                dn = ws.defined_names.get(name)
            except Exception:
                dn = None
            if dn is not None:
                break
    if dn is None:
        return ""
    try:
        for sheet_title, ref in dn.destinations:
            try:
                ws = wb_opxl[sheet_title]
            except KeyError:
                continue
            r1, r2, c1, c2 = parse_range_ref(ref)
            v = ws.cell(r1, c1).value
            if v is not None and str(v).strip():
                return str(v).strip()
    except Exception:
        pass
    return ""


def get_machine_info_opxl(wb_opxl) -> dict:
    """
    Read machine metadata.  Prefers named ranges (Customer, MachType, SerialNo,
    var_PLC); falls back to the legacy SUMMARY label-scan if a range is missing.
    """
    info = {
        "customer": _resolve_named_range(wb_opxl, "Customer"),
        "type":     _resolve_named_range(wb_opxl, "MachType"),
        "serial":   _resolve_named_range(wb_opxl, "SerialNo"),
        "plc":      "",
    }
    plc_raw = _resolve_named_range(wb_opxl, "var_PLC")
    if plc_raw:
        info["plc"] = VFD_MAP.get(plc_raw.upper(), plc_raw)

    # Legacy fallback: scan SUMMARY for any value we couldn't resolve above
    if not all((info["customer"], info["type"], info["serial"], info["plc"])):
        key_map = {"SERIAL NUMBER": "serial", "TYPE": "type", "CUSTOMER": "customer"}
        try:
            ws = wb_opxl["SUMMARY"]
            for row in ws.iter_rows(min_row=1, max_row=60, max_col=6):
                k = str(row[0].value or "").strip().upper()
                if k in key_map and not info[key_map[k]] and len(row) > 1 and row[1].value is not None:
                    info[key_map[k]] = str(row[1].value).strip()
                if k == "VAR_PLC" and not info["plc"] and len(row) > 1 and row[1].value is not None:
                    plc = str(row[1].value).strip().upper()
                    info["plc"] = VFD_MAP.get(plc, plc)
        except KeyError:
            pass
    return info


def _get_sheet_range(ws) -> tuple:
    """Return (r1, r2, c1, c2) from print area or used range."""
    max_row = ws.max_row or 1
    min_row = ws.min_row or 1
    min_col = ws.min_column or 1
    max_col = ws.max_column or 1
    pa = get_print_area(ws)
    if pa:
        try:
            r1, r2, c1, c2 = parse_range_ref(pa, max_row=max_row)
            return max(r1, min_row), min(r2, max_row), c1, c2
        except Exception as e:
            print(f"    Could not parse print area '{pa}': {e}, using used range")
    return min_row, max_row, min_col, max_col


def _find_subsection_rows(ws, r1: int, r2: int, c1: int, c2: int,
                          extra_keywords: set | None = None) -> list[tuple]:
    """
    Scan rows r1..r2 for sub-section dividers (full-width merged cells whose
    value is in SUB_SECTION_KEYWORDS or extra_keywords).
    Returns [(row, label), ...] sorted by row.
    """
    keywords = SUB_SECTION_KEYWORDS | ({k.upper() for k in extra_keywords} if extra_keywords else set())

    covered = set()
    for mr in ws.merged_cells.ranges:
        for rr in range(mr.min_row, mr.max_row + 1):
            for cc in range(mr.min_col, mr.max_col + 1):
                if rr != mr.min_row or cc != mr.min_col:
                    covered.add((rr, cc))

    dividers = []
    for r in range(r1, r2 + 1):
        non_empty = [c for c in range(c1, c2 + 1)
                     if (r, c) not in covered
                     and ws.cell(r, c).value is not None
                     and str(ws.cell(r, c).value).strip()]
        if len(non_empty) == 1:
            val = str(ws.cell(r, non_empty[0]).value or "").strip().upper()
            if val in keywords:
                dividers.append((r, val.title()))
    return dividers


def process_screens_opxl(ws, r1: int, r2: int, c1: int, c2: int,
                         assets_idx: dict, sheet_name: str,
                         assets_dir: Path = None) -> list[dict]:
    """
    Parse a SCREENS-type sheet where multiple HMI screens are laid out side-by-side.
    - Row r1: screen names in merged column groups
    - Row r1+1: screenshot images (one per group)
    - Remaining rows: OPTION | DESCRIPTION data filtered by col-A boolean
    Returns a single carousel HTML segment.
    """
    if assets_dir is None:
        assets_dir = Path(".")
    # ── Find column groups from merged cells in row r1 ─────────────────────
    groups = []  # [(col_start, col_end, screen_name)]
    for mr in ws.merged_cells.ranges:
        if (mr.min_row <= r1 <= mr.max_row
                and mr.max_col > mr.min_col
                and mr.min_col >= c1 and mr.max_col <= c2):
            val = str(ws.cell(mr.min_row, mr.min_col).value or "").strip()
            if val:
                groups.append((mr.min_col, mr.max_col, val))

    if not groups:
        # No merged groups — fall back to regular rendering with FALSE filter
        html = rows_to_html(ws, r1, r2, c1, c2,
                            sheet_name=sheet_name, assets_idx=assets_idx)
        return [{"type": "subsection", "label": "HMI Screens", "html": html}]

    groups.sort(key=lambda g: g[0])

    # ── Locate OPTION/DESCRIPTION header row within first 5 rows ──────────
    data_r1 = r1 + 2   # default: skip name row + image row
    option_col  = {}   # group_name → absolute column
    desc_col    = {}
    for r in range(r1 + 1, min(r1 + 6, r2 + 1)):
        row_vals = {c: str(ws.cell(r, c).value or "").strip().upper()
                    for c in range(c1, c2 + 1)}
        if any(v in ("OPTION", "DESCRIPTION", "DESC") for v in row_vals.values()):
            for gc1, gc2, gname in groups:
                for c in range(gc1, gc2 + 1):
                    v = row_vals.get(c, "")
                    if v == "OPTION" and gname not in option_col:
                        option_col[gname] = c
                    elif v in ("DESCRIPTION", "DESC") and gname not in desc_col:
                        desc_col[gname] = c
            data_r1 = r + 1
            break

    # ── Map screenshot images to column groups ─────────────────────────────
    # VBA ExportScreenImages writes assets/<SheetClean>/<CleanTitle>.png
    # where CleanTitle = _vba_clean_name(screen title).
    img_rows = {}   # group_name → src path relative to assets/ root
    sheet_clean = _vba_clean_name(sheet_name)
    for gc1, gc2, gname in groups:
        clean_title = _vba_clean_name(gname)
        # Preferred: subfolder named file from ExportScreenImages
        candidate = assets_dir / sheet_clean / (clean_title + ".png")
        if candidate.exists():
            img_rows[gname] = f"{sheet_clean}/{clean_title}.png"
            continue
        # Fallback: flat cell-address file from legacy ExportFlatImages
        for (ir, ic), fname in assets_idx.items():
            if r1 < ir < data_r1 and gc1 <= ic <= gc2 and gname not in img_rows:
                img_rows[gname] = fname

    # ── Build each screen's inner HTML (image + object/description table) ──
    slides = []  # list of (title, inner_html)
    for gc1, gc2, gname in groups:
        parts = []
        if gname in img_rows:
            parts.append(
                f'<div class="screen-img">'
                f'<img src="assets/{url_quote(img_rows[gname])}" alt="{gname}" '
                f'loading="lazy">'
                f'</div>'
            )
        opt_c  = option_col.get(gname, gc1)
        desc_c = desc_col.get(gname, gc1 + 1 if gc1 < gc2 else gc1)
        rows_out = []
        for r in range(data_r1, r2 + 1):
            if ws.cell(r, 1).value is False:
                continue
            opt_val  = format_cell_value(ws.cell(r, opt_c).value)
            desc_val = format_cell_value(ws.cell(r, desc_c).value)
            if not opt_val.strip() and not desc_val.strip():
                continue
            oe = opt_val.replace("&", "&amp;").replace("<", "&lt;")
            de = desc_val.replace("&", "&amp;").replace("<", "&lt;")
            rows_out.append(f"<tr><td>{oe}</td><td>{de}</td></tr>")
        if rows_out:
            parts.append(
                '<table class="data-table">'
                '<tr><th>Option</th><th>Description</th></tr>'
                + "".join(rows_out) + '</table>'
            )
        if parts:
            slides.append((gname, "".join(parts)))

    if not slides:
        return [{"type": "subsection", "label": "HMI Screens", "html": ""}]

    # ── Wrap as carousel: thumbnails + nav bar + slide stack ───────────────
    total = len(slides)
    thumbs_html = "".join(
        f'<button class="hmi-thumb{" active" if i == 0 else ""}" '
        f'type="button" data-idx="{i}">{title}</button>'
        for i, (title, _) in enumerate(slides)
    )
    slides_html = "".join(
        f'<div class="hmi-slide{" active" if i == 0 else ""}" data-idx="{i}">'
        f'<h3 class="hmi-slide-title">{title}</h3>{inner}</div>'
        for i, (title, inner) in enumerate(slides)
    )
    carousel_html = (
        '<div class="hmi-carousel" data-current="0">'
        '<div class="hmi-carousel-bar">'
        '<button class="hmi-prev" type="button" aria-label="Previous screen">&lsaquo;</button>'
        f'<span class="hmi-counter"><span class="hmi-cur">1</span> / <span class="hmi-total">{total}</span></span>'
        '<button class="hmi-next" type="button" aria-label="Next screen">&rsaquo;</button>'
        '</div>'
        f'<div class="hmi-thumbs">{thumbs_html}</div>'
        f'<div class="hmi-slides">{slides_html}</div>'
        '</div>'
    )
    return [{"type": "html", "content": carousel_html}]


def process_sheet_opxl(sheet_name: str, wb_opxl, assets_dir: Path,
                       sheet_type: str = "") -> list[dict]:
    """
    Render one sheet to HTML segments using only openpyxl.
    - sheet_type='module': splits at SUB_SECTION_KEYWORDS dividers → subsection segments
    - Named tables tbl_BOM_* / tbl_EEBOM → rendered as BOM tables appended to segments
    - Images loaded from assets/{sheet_clean}/{section_clean}/ subfolders (VBA A1-note scheme)
    """
    try:
        ws = wb_opxl[sheet_name]
    except KeyError:
        print(f"    Sheet not found in workbook: '{sheet_name}'")
        return []

    # A1 note metadata (INCLUDE / TYPE / SECTIONS) written by VBA
    meta = _read_sheet_meta(ws)
    if meta["INCLUDE"] == "FALSE":
        print(f"    [{sheet_name}] skipped (INCLUDE=FALSE)")
        return []
    # A1-note TYPE overrides tbl_PRINT TYPE when present and not REGULAR
    if meta["TYPE"] not in ("REGULAR", "") and sheet_type == "":
        sheet_type = meta["TYPE"].lower()
    # Defensive fallback: infer type from sheet name (SCREENS in particular
    # doesn't always carry a full A1 note — INCLUDE=TRUE; alone is allowed)
    if not sheet_type:
        sheet_type = _sheet_type_from_name(sheet_name)
    meta_sections = [s.strip() for s in meta["SECTIONS"].split(",") if s.strip()]

    r1, r2, c1, c2 = _get_sheet_range(ws)
    # Flat-asset index kept for backward-compat; new exports use subfolders
    assets_idx = _build_assets_index(assets_dir, sheet_name)

    print(f"    [{sheet_name}] print={r1}:{r2} cols {c1}:{c2} | "
          f"flat-assets={len(assets_idx)} | meta={dict(meta)}")

    # ── Named BOM tables (tbl_BOM_* and tbl_EEBOM) ────────────────────────────
    bom_table_html = []
    for tname, tobj in ws.tables.items():
        is_bom_tbl = (tname.upper().startswith(BOM_TABLE_PREFIX.upper())
                      or tname.upper() in {n.upper() for n in BOM_TABLE_NAMES})
        if is_bom_tbl:
            try:
                tr1, tr2, tc1, tc2 = parse_range_ref(tobj.ref, max_row=ws.max_row or 1)
                html = rows_to_html(ws, tr1, tr2, tc1, tc2,
                                    sheet_name=sheet_name, assets_idx=assets_idx,
                                    sheet_type=sheet_type, force_bom=True)
                if html:
                    bom_table_html.append(html)
            except Exception as e:
                print(f"    Could not render BOM table {tname}: {e}")

    # Discover all section images once (filesystem-driven, robust to name variants)
    all_images = load_module_images(assets_dir, sheet_name)
    print(f"    [{sheet_name}] image sections found: {sorted(all_images)}")

    # ── Module / maintenance sheets: split by sub-section dividers ───────────
    if sheet_type in ("module", "maintenance"):
        dividers = _find_subsection_rows(ws, r1, r2, c1, c2,
                                         extra_keywords=set(meta_sections) if meta_sections else None)
        segments = []
        if not dividers:
            img_html = all_images.get("Overview", "")
            tbl_html = rows_to_html(ws, r1, r2, c1, c2,
                                    sheet_name=sheet_name, assets_idx=assets_idx,
                                    sheet_type=sheet_type)
            combined = img_html + (tbl_html or "")
            if combined:
                segments.append({"type": "subsection", "label": "Overview", "html": combined})
        else:
            # Content before first divider → Overview
            first_div_row = dividers[0][0]
            if first_div_row > r1:
                img_html = all_images.get("Overview", "")
                tbl_html = rows_to_html(ws, r1, first_div_row - 1, c1, c2,
                                        sheet_name=sheet_name, assets_idx=assets_idx,
                                        sheet_type=sheet_type)
                combined = img_html + (tbl_html or "")
                if combined:
                    segments.append({"type": "subsection", "label": "Overview", "html": combined})
            # Each labelled divider section
            for idx, (div_row, label) in enumerate(dividers):
                next_start = dividers[idx + 1][0] if idx + 1 < len(dividers) else r2 + 1
                img_html = all_images.get(label, "")   # label is already Title-cased
                tbl_html = rows_to_html(ws, div_row + 1, next_start - 1, c1, c2,
                                        sheet_name=sheet_name, assets_idx=assets_idx,
                                        sheet_type=sheet_type)
                combined = img_html + (tbl_html or "")
                if combined or label == "Bom":
                    segments.append({"type": "subsection", "label": label, "html": combined})
        # Append named BOM tables
        for bhtml in bom_table_html:
            existing_bom = next((s for s in segments if s["label"] == "Bom"), None)
            if existing_bom:
                existing_bom["html"] += bhtml
            else:
                segments.append({"type": "subsection", "label": "Bom", "html": bhtml})
        return segments

    # ── Screens sheet: column-group carousel ──────────────────────────────────
    if sheet_type == "screens":
        return process_screens_opxl(ws, r1, r2, c1, c2, assets_idx, sheet_name,
                                    assets_dir=assets_dir)

    # ── Regular / settings sheet ──────────────────────────────────────────────
    # Prepend images from all discovered section folders
    all_img_html = "".join(all_images.values())
    tbl_html = rows_to_html(ws, r1, r2, c1, c2,
                            sheet_name=sheet_name, assets_idx=assets_idx,
                            sheet_type=sheet_type)
    combined = all_img_html + (tbl_html or "")
    segs = [{"type": "html", "content": combined}] if combined else []
    for bhtml in bom_table_html:
        segs.append({"type": "html", "content": bhtml})
    return segs


def _rgb_from_color(color_obj) -> str | None:
    """Extract a #RRGGBB hex string from an openpyxl Color object, or None."""
    if color_obj is None:
        return None
    try:
        if color_obj.type == "rgb":
            rgb = color_obj.rgb  # 'FFRRGGBB' or 'RRGGBB'
            if rgb and rgb not in ("00000000", "FF000000", "000000"):
                hex6 = rgb[-6:]  # last 6 chars = RRGGBB
                if hex6 != "000000":
                    return f"#{hex6}"
    except Exception:
        pass
    return None


def _is_header_row(ws, r: int, c1: int, c2: int, covered: set) -> bool:
    """
    Return True if this row is sheet-header boilerplate that should be skipped.
    Checks column A (always) and c1, and also catches single-value title rows
    that span the full width (merged title cells).
    """
    # Check column 1 (A) and c1 for known keywords
    for check_col in {1, c1}:
        val = str(ws.cell(r, check_col).value or "").strip().upper().rstrip(":")
        if val in HEADER_KEYWORDS:
            return True

    # Also skip rows where only one cell has a value and it spans the whole width
    # (these are sheet-title merged cells like "CLEARPRINT-24L INFEED CONVEYOR")
    # BUT do NOT skip sub-section keywords — they become tab dividers.
    non_empty = []
    for c in range(c1, c2 + 1):
        if (r, c) not in covered:
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                non_empty.append(c)
    if len(non_empty) == 1:
        val = str(ws.cell(r, non_empty[0]).value or "").strip().upper()
        if val in SUB_SECTION_KEYWORDS:
            return False   # sub-section divider — keep
        if val == "PICTURE":
            return False   # VBA picture-in-cell marker — keep as image row
        if len(val) > 80:
            return False   # long descriptive text — keep as content
        for mr in ws.merged_cells.ranges:
            if (mr.min_row == r and mr.min_col == non_empty[0]
                    and mr.max_col >= c2 - 1):
                return True  # short text spanning full width → sheet title banner
    return False


def rows_to_html(ws, r1: int, r2: int, c1: int, c2: int,
                 sheet_name: str = "", assets_dir: Path = None,
                 assets_idx: dict = None,
                 sheet_type: str = "", force_bom: bool = False) -> str:
    """
    Render rows r1..r2, cols c1..c2 from an openpyxl worksheet as an HTML table.
    - assets_idx: pre-built {(row,col): filename} from _build_assets_index (preferred)
    - assets_dir: fallback assets directory if assets_idx not provided
    - force_bom: treat as BOM table regardless of column header detection
    """
    if r1 > r2:
        return ""

    # Build merged-cell maps
    merged_spans: dict[tuple, tuple] = {}
    covered: set[tuple] = set()
    for mr in ws.merged_cells.ranges:
        if mr.max_row < r1 or mr.min_row > r2 or mr.max_col < c1 or mr.min_col > c2:
            continue
        rowspan = mr.max_row - mr.min_row + 1
        colspan = mr.max_col - mr.min_col + 1
        merged_spans[(mr.min_row, mr.min_col)] = (rowspan, colspan)
        for rr in range(mr.min_row, mr.max_row + 1):
            for cc in range(mr.min_col, mr.max_col + 1):
                if rr != mr.min_row or cc != mr.min_col:
                    covered.add((rr, cc))

    # Build assets index if not passed in
    if assets_idx is None:
        assets_idx = _build_assets_index(assets_dir, sheet_name) if assets_dir else {}

    # ── Pre-scan: images in assets_idx that may be at col < c1 ──────────
    # VBA exports from the merged cell's actual top-left, which may be col A
    # even when the print area starts at col B or later.
    pre_area_imgs = {}  # row → fname for images outside print area column range
    for (ir, ic), fname in assets_idx.items():
        if r1 <= ir <= r2 and ic < c1:
            pre_area_imgs[ir] = fname

    # ── Detect BOM table: forced or first row has BOM column header ──────
    first_row_vals = []
    # Scan up to 6 rows to find BOM column headers (image rows may appear first)
    # Part-number priority: PART # / POLYPACK PN > ITEM # / NUMBER
    PN_PRIMARY   = {"PART #", "PART NUMBER", "P/N", "PN", "POLYPACK PN", "POLYPACK P/N"}
    PN_FALLBACK  = {"ITEM #", "ITEM NO", "NUMBER"}
    first_row_vals = []
    bom_header_row = r1
    for scan_r in range(r1, min(r1 + 6, r2 + 1)):
        rv = []
        for c in range(c1, min(c1 + 14, c2 + 1)):
            rv.append(str(ws.cell(scan_r, c).value or "").strip().upper())
        if any(v in BOM_COL_KEYWORDS for v in rv):
            first_row_vals = rv
            bom_header_row = scan_r
            break
    if not first_row_vals:
        for c in range(c1, min(c1 + 14, c2 + 1)):
            first_row_vals.append(str(ws.cell(r1, c).value or "").strip().upper())

    is_bom = force_bom or any(v in BOM_COL_KEYWORDS for v in first_row_vals)
    # Prefer a true part-number column over an item-number column
    pn_idx = next((i for i, v in enumerate(first_row_vals) if v in PN_PRIMARY), None)
    if pn_idx is None:
        pn_idx = next((i for i, v in enumerate(first_row_vals) if v in PN_FALLBACK), None)
    desc_idx   = next((i for i, v in enumerate(first_row_vals) if "DESC" in v), None)
    config_idx = next((i for i, v in enumerate(first_row_vals)
                       if "CONFIG" in v or "CONF" in v or "SUPPLIER PN" in v), None)
    qty_idx    = next((i for i, v in enumerate(first_row_vals) if v in ("QTY", "QUANTITY")), None)

    # ── Detect PM "IMG" icon column ──────────────────────────────────────
    img_col_idx = next((i for i, v in enumerate(first_row_vals) if v == "IMG"), None)

    # ── Detect PROCESS column beyond the print area (PM icon lookup) ──────
    # IMG column formula =XLOOKUP([@PROCESS],...) resolves to an image object
    # openpyxl can't read image objects, so we read the source PROCESS column
    # (which is outside the print range) and map its value to an HTML icon.
    process_col_abs = None  # absolute column index in the worksheet
    if img_col_idx is not None:
        scan_limit = min(c2 + 20, (ws.max_column or c2 + 20))
        for extra_c in range(c2 + 1, scan_limit + 1):
            hdr = str(ws.cell(r1, extra_c).value or "").strip().upper()
            if hdr == "PROCESS":
                process_col_abs = extra_c
                break

    # ── Detect boolean-only columns (include/exclude checkboxes) ─────────
    # Scan a sample of data rows to find columns that are all True/False
    bool_cols = set()
    sample_rows = list(range(r1 + 1, min(r1 + 10, r2 + 1)))
    if sample_rows:
        for col_i, col_v in enumerate(first_row_vals):
            c_abs = c1 + col_i
            vals = [ws.cell(sr, c_abs).value for sr in sample_rows]
            non_none = [v for v in vals if v is not None]
            if non_none and all(isinstance(v, bool) for v in non_none):
                bool_cols.add(col_i)

    is_settings = sheet_type == "settings"
    ncols_visible = c2 - c1 + 1 - len(bool_cols)  # for colspan in image rows

    rows_html = []
    first_data_row = True
    for r in range(r1, r2 + 1):
        if _is_header_row(ws, r, c1, c2, covered):
            continue

        # ── Skip rows where any boolean column = False ──────────────────
        if bool_cols:
            row_excluded = any(
                ws.cell(r, c1 + bi).value is False
                for bi in bool_cols
            )
            if row_excluded:
                continue

        # ── Skip Picture-placeholder rows with no flat-asset image ───────
        # Images now live in section subfolders (injected above the table);
        # without this, Picture rows render as blank table rows.
        _pic_anchor = next(
            (chk for chk in range(c1, c2 + 1)
             if (r, chk) not in covered
             and str(ws.cell(r, chk).value or "").strip().upper() == "PICTURE"),
            None,
        )
        if _pic_anchor is not None:
            _has_flat = (
                any(assets_idx.get((r, chk)) for chk in range(c1, c2 + 1))
                or r in pre_area_imgs
            )
            if not _has_flat:
                continue

        # ── Inject images from columns before the print area ───────────
        if r in pre_area_imgs:
            nc = max(ncols_visible, 1)
            rows_html.append(
                f'<tr><td colspan="{nc}" style="padding:0;text-align:center">'
                f'<img src="assets/{url_quote(pre_area_imgs[r])}" alt="{sheet_name}" '
                f'style="max-width:100%;height:auto;display:block;">'
                f'</td></tr>'
            )
            # If the row has no other data in the print area, skip the cell loop
            _SKIP_VALS = {None, "", "Picture", "PICTURE"}
            non_img_vals = [
                ws.cell(r, c).value for c in range(c1, c2 + 1)
                if (r, c) not in covered
                and ws.cell(r, c).value not in _SKIP_VALS
            ]
            if not non_img_vals:
                continue

        cells_html = []
        row_vals = []

        for c in range(c1, c2 + 1):
            if (r, c) in covered:
                continue

            col_index = c - c1  # 0-based
            # Skip boolean-only columns
            if col_index in bool_cols:
                continue

            cell = ws.cell(r, c)

            # ── Image lookup ────────────────────────────────────────────
            img_tag = None
            img_fname = assets_idx.get((r, c))
            if img_fname:
                img_tag = (f'<img src="assets/{url_quote(img_fname)}" '
                           f'alt="{sheet_name}" '
                           f'style="max-width:100%;height:auto;display:block;">')

            raw_val = format_cell_value(cell.value)
            # Suppress "Picture" marker text — it's an image placeholder, not content
            if raw_val.strip().upper() == "PICTURE":
                raw_val = ""
            row_vals.append(raw_val)
            text = img_tag if img_tag else (raw_val or "&nbsp;")

            # ── PM IMG column → icon ─────────────────────────────────────
            if img_col_idx is not None and col_index == img_col_idx and not img_tag:
                # Prefer the PROCESS column (outside print area) as the lookup key;
                # fall back to the IMG cell value if PROCESS isn't available.
                if process_col_abs is not None:
                    proc_val = str(ws.cell(r, process_col_abs).value or "").strip().upper()
                else:
                    proc_val = raw_val.strip().upper()
                text = PM_IMG_ICONS.get(proc_val, raw_val or "&nbsp;")

            # ── Settings mode: render value cells as inputs ─────────────
            if is_settings and not first_data_row and col_index >= 1 and not img_tag:
                esc = raw_val.replace('"', '&quot;')
                text = (f'<div class="sv-wrap">'
                        f'<input class="sv-input" type="text" value="{esc}" '
                        f'data-orig="{esc}">'
                        f'<span class="sv-dot"></span>'
                        f'</div>')

            # ── Style ──────────────────────────────────────────────────────
            style: list[str] = []
            try:
                fill = cell.fill
                if fill and fill.patternType == "solid":
                    bg = _rgb_from_color(fill.fgColor)
                    if bg:
                        style.append(f"background:{bg}")
            except Exception:
                pass
            try:
                font = cell.font
                if font:
                    if font.bold:
                        style.append("font-weight:bold")
                    fc = _rgb_from_color(font.color)
                    # Skip near-white font colors — page is light-themed
                    if fc and not _is_light_color(fc):
                        style.append(f"color:{fc}")
                    if font.size and font.size > 0:
                        style.append(f"font-size:{int(font.size)}pt")
            except Exception:
                pass
            try:
                al = cell.alignment
                if al:
                    if al.horizontal in ("center", "centerContinuous"):
                        style.append("text-align:center")
                    elif al.horizontal == "right":
                        style.append("text-align:right")
                    if al.wrap_text:
                        style.append("white-space:pre-wrap")
            except Exception:
                pass

            style_attr = f' style="{";".join(style)}"' if style else ""
            span    = merged_spans.get((r, c), (1, 1))
            rs_attr = f' rowspan="{span[0]}"' if span[0] > 1 else ""
            cs_attr = f' colspan="{span[1]}"' if span[1] > 1 else ""
            cells_html.append(f"<td{style_attr}{rs_attr}{cs_attr}>{text}</td>")

        if cells_html:
            # BOM add-to-quote button — only on data rows (after the header row)
            # Read directly by absolute column to avoid row_vals index drift from
            # covered / bool-skipped cells.
            if is_bom and r > bom_header_row:
                def _bom_val(idx):
                    if idx is None:
                        return ""
                    return format_cell_value(ws.cell(r, c1 + idx).value)
                pn   = _bom_val(pn_idx)
                desc = _bom_val(desc_idx)
                cfg  = _bom_val(config_idx)
                qty  = _bom_val(qty_idx)
                if pn.strip():
                    epn   = pn.replace('"', '&quot;')
                    edesc = desc.replace('"', '&quot;')
                    ecfg  = cfg.replace('"', '&quot;')
                    eqty  = qty.replace('"', '&quot;')
                    emod  = sheet_name.replace('"', '')
                    btn = (f'<td class="bom-btn-cell">'
                           f'<button class="add-quote-btn" '
                           f'data-pn="{epn}" data-desc="{edesc}" '
                           f'data-cfg="{ecfg}" data-qty="{eqty}" '
                           f'data-module="{emod}">+</button></td>')
                    cells_html.append(btn)

            rows_html.append(f"<tr>{''.join(cells_html)}</tr>")
            first_data_row = False

    if not rows_html:
        return ""

    tbl_class = "data-table settings-table" if is_settings else "data-table"
    if is_bom:
        tbl_class += " bom-table"
    table = f'<table class="{tbl_class}">' + "".join(rows_html) + "</table>"
    if len(rows_html) > 25:
        table = '<div class="table-scroll">' + table + "</div>"
    return table


# ---------------------------------------------------------------------------
# win32com helpers
# ---------------------------------------------------------------------------

def open_workbook(xlsx_path: Path):
    """Open Excel invisibly. Returns (xl_app, workbook)."""
    pythoncom.CoInitialize()
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible       = False
    xl.DisplayAlerts = False
    wb = xl.Workbooks.Open(str(xlsx_path.resolve()))
    return xl, wb


def export_range_png(wb_com, sheet_name: str, range_ref: str, out_path: Path) -> bool:
    """
    Export a named range as a PNG via CopyPicture → temp chart → Export.
    Captures everything: cell images, overlaid shapes, callouts, arrows.
    Excel must be briefly visible for CopyPicture to work reliably.
    """
    xl_app = wb_com.Application
    try:
        xl_app.Visible = True          # CopyPicture requires a visible window
        xl_app.ScreenUpdating = False  # but suppress screen flicker
        ws  = wb_com.Sheets(sheet_name)
        ws.Activate()
        rng = ws.Range(range_ref)
        rng.CopyPicture(Appearance=2, Format=2)  # xlPrinter=2, xlBitmap=2
        co  = ws.ChartObjects().Add(0, 0, rng.Width, rng.Height)
        co.Chart.Paste()
        co.Chart.Export(str(out_path.resolve()))
        co.Delete()
        return True
    except Exception as e:
        print(f"    PNG export failed ({sheet_name} {range_ref}): {e}")
        return False
    finally:
        try:
            xl_app.Visible = False
            xl_app.ScreenUpdating = True
        except Exception:
            pass


def get_com_print_area(wb_com, sheet_name: str) -> str | None:
    """Get the print area reference from Excel COM (more reliable than openpyxl)."""
    try:
        pa = wb_com.Sheets(sheet_name).PageSetup.PrintArea
        return pa if pa else None
    except Exception:
        return None


def read_print_table(wb_com) -> list[str]:
    """Read tbl_PRINT from ADMIN sheet via COM. Returns ordered sheet name list."""
    try:
        ws = wb_com.Sheets("ADMIN")
    except Exception:
        raise RuntimeError("ADMIN sheet not found. Please add an ADMIN sheet with tbl_PRINT.")

    # Try ListObjects (named Excel Table)
    for lo in ws.ListObjects:
        if lo.Name == "tbl_PRINT":
            headers = []
            for c in range(1, lo.HeaderRowRange.Columns.Count + 1):
                headers.append(str(lo.HeaderRowRange.Cells(1, c).Value or "").strip().upper())
            name_col = headers.index("NAME") + 1 if "NAME" in headers else 2
            names = []
            if lo.DataBodyRange is not None:
                for r in range(1, lo.DataBodyRange.Rows.Count + 1):
                    v = lo.DataBodyRange.Cells(r, name_col).Value
                    if v:
                        names.append(str(v).strip())
            return names

    # Fallback: scan for NAME header
    for row in range(1, 15):
        for col in range(1, 10):
            if str(ws.Cells(row, col).Value or "").strip().upper() == "NAME":
                names, r = [], row + 1
                while True:
                    v = ws.Cells(r, col).Value
                    if not v:
                        break
                    names.append(str(v).strip())
                    r += 1
                return names

    raise RuntimeError("tbl_PRINT not found in ADMIN sheet.")


def get_machine_info(wb_com) -> dict:
    """Read serial, type, customer from SUMMARY sheet."""
    info = {"serial": "", "type": "", "customer": ""}
    key_map = {"SERIAL NUMBER": "serial", "TYPE": "type", "CUSTOMER": "customer"}
    try:
        ws = wb_com.Sheets("SUMMARY")
        for row in range(1, 30):
            k = str(ws.Cells(row, 1).Value or "").strip().upper()
            if k in key_map:
                v = ws.Cells(row, 2).Value
                if v is not None:
                    info[key_map[k]] = str(v).strip()
    except Exception as e:
        print(f"  Warning: could not read machine info: {e}")
    return info


# ---------------------------------------------------------------------------
# Sheet processor — builds a list of segments
# ---------------------------------------------------------------------------

def process_sheet(
    sheet_name: str,
    wb_com,
    wb_opxl,
    assets_dir: Path,
) -> list[dict]:
    """
    Returns an ordered list of segments for one sheet:
      {type: 'html', content: str}
      {type: 'img',  file: str, alt: str}   (file is relative to assets/)
    """

    # ── Print area ───────────────────────────────────────────────────────────
    pa_ref = get_com_print_area(wb_com, sheet_name)
    if not pa_ref:
        # Fall back to openpyxl
        try:
            ws_opxl_tmp = wb_opxl[sheet_name]
            pa_ref = get_print_area(ws_opxl_tmp)
        except Exception:
            pass

    # Get used range from COM for fallback / clipping column-only refs
    used_r1, used_r2, used_c1, used_c2 = 1, 1, 1, 1
    try:
        ur = wb_com.Sheets(sheet_name).UsedRange
        used_r1 = ur.Row
        used_r2 = ur.Row + ur.Rows.Count - 1
        used_c1 = ur.Column
        used_c2 = ur.Column + ur.Columns.Count - 1
    except Exception:
        pass

    if not pa_ref:
        # No print area set — use full used range
        print(f"    No print area for '{sheet_name}', using used range")
        pa_r1, pa_r2, pa_c1, pa_c2 = used_r1, used_r2, used_c1, used_c2
    else:
        try:
            pa_r1, pa_r2, pa_c1, pa_c2 = parse_range_ref(pa_ref, max_row=used_r2)
            # Clip column-only refs (e.g. $A:$G where pa_r2 was set to max_row)
            pa_r1 = max(pa_r1, used_r1)
            pa_r2 = min(pa_r2, used_r2)
        except Exception as e:
            print(f"    Could not parse print area '{pa_ref}': {e}")
            return []

    # ── Named ranges on this sheet ───────────────────────────────────────────
    try:
        ws_opxl = wb_opxl[sheet_name]
    except KeyError:
        print(f"    Sheet not found in openpyxl: '{sheet_name}'")
        return []

    img_ranges = get_named_ranges(wb_opxl, sheet_name, "img_")
    hdr_ranges = get_named_ranges(wb_opxl, sheet_name, "hdr_")

    # Build a unified sorted event list for all special row regions
    # type 'img'  → export as PNG
    # type 'skip' → omit entirely (e.g. hdr_* sheet headers not needed on web)
    events = sorted(
        [dict(e, type="img")  for e in img_ranges] +
        [dict(e, type="skip") for e in hdr_ranges],
        key=lambda x: x["r1"]
    )

    # ── Walk print area top-to-bottom ────────────────────────────────────────
    segments: list[dict] = []
    current_row = pa_r1

    for ev in events:
        er1 = max(ev["r1"], pa_r1)
        er2 = min(ev["r2"], pa_r2)
        if er1 > er2 or er1 < current_row:
            continue

        # HTML for rows between current position and this event
        if current_row < er1:
            html = rows_to_html(ws_opxl, current_row, er1 - 1, pa_c1, pa_c2)
            if html:
                segments.append({"type": "html", "content": html})

        if ev["type"] == "img":
            safe_name = re.sub(r"[^a-z0-9]+", "_", ev["name"].lower()).strip("_")
            img_file  = f"{safe_name}.png"
            out_path  = assets_dir / img_file
            print(f"    → PNG: {ev['name']} ({ev['ref']})", end=" ", flush=True)
            ok = export_range_png(wb_com, sheet_name, ev["ref"], out_path)
            print("OK" if ok else "FAILED")
            segments.append({"type": "img", "file": img_file if ok else None, "alt": ev["name"]})
        # type == 'skip': rows silently omitted

        current_row = er2 + 1

    # HTML for any remaining rows after last event (or all rows if no events)
    if current_row <= pa_r2:
        html = rows_to_html(ws_opxl, current_row, pa_r2, pa_c1, pa_c2)
        if html:
            segments.append({"type": "html", "content": html})

    return segments


# ---------------------------------------------------------------------------
# HTML assembly
# ---------------------------------------------------------------------------

CSS = r"""
@import url('https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;600;700&family=Barlow:wght@400;500;600&family=DM+Mono:wght@400;500&display=swap');

:root {
  --sw:       240px;
  --red:      #cc0000;
  --red-dim:  rgba(204,0,0,0.09);
  --black:    #1a1a1a;
  --bg:       #f2f3f5;
  --surface:  #ffffff;
  --surface2: #f8f8fa;
  --border:   #e0e0e4;
  --border2:  #d0d0d6;
  --txt:      #1a1a1a;
  --txt-muted:#5a5a6a;
  --txt-dim:  #9a9aaa;
  --row-alt:  #f6f6f9;
  --sans:     'Barlow', system-ui, Arial, sans-serif;
  --cond:     'Barlow Condensed', Arial Narrow, Arial, sans-serif;
  --mono:     'DM Mono', 'Courier New', monospace;
  --shadow:   0 1px 3px rgba(0,0,0,.08), 0 2px 8px rgba(0,0,0,.05);
}
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
body {
  font-family: var(--sans); font-size: 14px;
  background: var(--bg); color: var(--txt);
  display: flex; flex-direction: column; min-height: 100vh;
}

/* ── Sidebar ─────────────────────────────────────────────────────────── */
/* ── Layout shell ────────────────────────────────────────────────────── */
#topbar {
  height: 48px; background: var(--black);
  border-bottom: 3px solid var(--red);
  display: flex; align-items: center; padding: 0 20px; gap: 16px;
  position: sticky; top: 0; z-index: 200; flex-shrink: 0;
}
.tb-brand {
  font-family: var(--cond); font-size: 17px; font-weight: 700;
  letter-spacing: .1em; text-transform: uppercase; color: #fff;
}
.tb-brand span { color: var(--red); }
.tb-divider { width: 1px; height: 18px; background: rgba(255,255,255,.18); }
.tb-machine { font-family: var(--cond); font-size: 14px; font-weight: 600; color: #ddd; letter-spacing: .04em; }
.tb-sn {
  font-family: var(--mono); font-size: 10px; letter-spacing: .1em;
  color: var(--red); background: rgba(204,0,0,.12);
  border: 1px solid rgba(204,0,0,.3); padding: 2px 8px; margin-left: 6px;
}
.tb-customer { font-family: var(--mono); font-size: 11px; color: #888; margin-left: auto; }

#app { display: flex; flex: 1; min-height: 0; }

/* ── Sidebar ─────────────────────────────────────────────────────────── */
#sidebar {
  width: var(--sw); min-width: var(--sw);
  background: var(--black);
  height: calc(100vh - 48px); position: sticky; top: 48px;
  overflow-y: auto; flex-shrink: 0;
  border-right: 1px solid rgba(255,255,255,.07);
}
.sb-section-label {
  font-family: var(--mono); font-size: 9px; letter-spacing: .16em;
  text-transform: uppercase; color: rgba(255,255,255,.25);
  padding: 16px 16px 6px; display: block;
}
.sb-section-label:first-child { padding-top: 14px; }

/* ── Sidebar group nav ────────────────────────────────────────────── */
.grp-nav { }
.grp-nav-hdr {
  display: flex; align-items: center;
  padding: 12px 16px 6px;
  font-family: var(--mono); font-size: 9px; letter-spacing: .16em;
  text-transform: uppercase; color: rgba(255,255,255,.35);
  cursor: pointer; user-select: none;
  transition: color .1s;
}
.grp-nav-hdr:hover { color: rgba(255,255,255,.6); }
.grp-nav-hdr::after { content: "\25BE"; margin-left: auto; color: var(--red); font-size: 10px; }
.grp-nav.nav-closed .grp-nav-hdr::after { content: "\25B8"; }
.grp-nav-sub { border-left: 2px solid rgba(204,0,0,.25); margin-left: 14px; }
.grp-nav.nav-closed .grp-nav-sub { display: none; }
.grp-nav-sub a {
  display: flex; align-items: center; gap: 8px;
  padding: 6px 14px;
  font-family: var(--cond); font-size: 12px; font-weight: 500;
  letter-spacing: .03em; text-transform: none;
  color: rgba(255,255,255,.45); text-decoration: none;
  border-left: none;
  transition: color .1s, background .1s;
}
.grp-nav-sub a:hover { color: #fff; background: rgba(255,255,255,.05); }
.grp-nav-sub a.active { color: #fff; background: rgba(204,0,0,.1); }
/* Module sub-list within a group nav */
.grp-nav-sub .mod-nav-label {
  padding: 6px 14px 2px;
  font-family: var(--mono); font-size: 8px; letter-spacing: .15em;
  text-transform: uppercase; color: rgba(255,255,255,.2);
  display: block;
}
.grp-nav-sub .mod-nav-list { border-left: 2px solid rgba(204,0,0,.15); margin-left: 14px; }
.grp-nav-sub .mod-nav-list a { font-size: 11px; padding: 5px 12px; }

/* ── Main content ────────────────────────────────────────────────────── */
#main { flex: 1; padding: 24px 32px; overflow-x: auto; min-width: 0; }

/* ── Machine info grid (top of page) ────────────────────────────────── */
.machine-info-grid {
  display: grid; grid-template-columns: repeat(3, 1fr);
  gap: 1px; background: var(--border);
  border: 1px solid var(--border);
  margin-bottom: 24px;
}
.mi-cell {
  background: var(--surface); padding: 14px 18px;
}
.mi-label {
  font-family: var(--mono); font-size: 9px; letter-spacing: .14em;
  text-transform: uppercase; color: var(--txt-dim); margin-bottom: 4px;
}
.mi-value {
  font-family: var(--cond); font-size: 20px; font-weight: 600;
  letter-spacing: .03em; color: var(--txt);
}
.mi-value.accent { color: var(--red); }
.mi-value.small { font-size: 14px; font-family: var(--mono); font-weight: 400; }

/* ── Group accordion ─────────────────────────────────────────────────── */
.grp-section {
  border: 1px solid var(--border);
  margin-bottom: 20px; overflow: hidden;
  box-shadow: var(--shadow);
}
.grp-section-hdr {
  display: flex; align-items: center;
  padding: 13px 18px;
  background: var(--surface2);
  border-left: 4px solid var(--red);
  cursor: pointer; user-select: none;
  transition: background .1s;
}
.grp-section-hdr:hover { background: #1e1e1e; }
.grp-section-hdr-title {
  font-family: var(--cond); font-size: 14px; font-weight: 800;
  letter-spacing: .12em; text-transform: uppercase; color: var(--txt);
  flex: 1;
}
.grp-section-hdr::after {
  content: "\25BE"; color: var(--red); font-size: 14px; margin-left: 10px;
}
.grp-section.collapsed .grp-section-hdr::after { content: "\25B8"; }
.grp-body { display: block; }
.grp-section.collapsed .grp-body { display: none; }

/* ── Sheet panel accordion ───────────────────────────────────────────── */
.sheet-panel {
  border-top: 1px solid var(--border);
  overflow: hidden;
  background: var(--surface);
}
.sheet-panel:first-child { border-top: none; }
.sheet-panel-hdr {
  display: flex; align-items: center;
  padding: 10px 18px;
  background: var(--surface);
  border-left: 3px solid rgba(204,0,0,.4);
  cursor: pointer; user-select: none;
  transition: background .1s;
}
.sheet-panel-hdr:hover { background: var(--surface2); }
.sheet-panel-hdr-title {
  font-family: var(--cond); font-size: 13px; font-weight: 700;
  letter-spacing: .08em; text-transform: uppercase; color: var(--txt-muted);
  flex: 1;
}
.sheet-panel-hdr-badge {
  font-family: var(--mono); font-size: 9px; letter-spacing: .1em;
  color: var(--txt-dim); background: var(--bg);
  border: 1px solid var(--border); padding: 2px 7px; margin-left: 8px;
}
.sheet-panel-hdr::after {
  content: "\25BE"; color: rgba(204,0,0,.6); font-size: 12px; margin-left: 10px;
}
.sheet-panel.collapsed .sheet-panel-hdr::after { content: "\25B8"; }
.sheet-panel-body { padding: 16px; }
.sheet-panel.collapsed .sheet-panel-body { display: none; }
/* Module panel: no padding — tab bar sits flush */
.sheet-panel.module-panel .sheet-panel-body { padding: 0; }

/* ── Sub-section header (dividers inside sheet content) ──────────────── */
.section-sub-hdr {
  display: flex; align-items: center;
  padding: 8px 16px;
  background: var(--surface2);
  border-bottom: 1px solid var(--border);
  border-left: 3px solid rgba(204,0,0,.4);
  font-family: var(--cond); font-size: 12px; font-weight: 600;
  letter-spacing: .08em; text-transform: uppercase; color: var(--txt-muted);
}

/* ── Images ──────────────────────────────────────────────────────────── */
.sheet-panel-body img, .sub-pane img {
  max-width: 100%; height: auto; display: block; margin: .5rem 0;
  border: 1px solid var(--border);
}

/* ── Tables ──────────────────────────────────────────────────────────── */
.data-table, .sheet-table {
  border-collapse: collapse; width: 100%;
  font-size: 12.5px; font-family: var(--sans);
  margin: 0;
}
.data-table th, .sheet-table th,
.data-table tr:first-child td, .sheet-table tr:first-child td {
  font-family: var(--mono); font-size: 9px; letter-spacing: .12em;
  text-transform: uppercase; color: var(--txt-dim);
  background: var(--surface2); padding: 8px 10px;
  border-bottom: 1px solid var(--border2);
  border-right: 1px solid var(--border);
}
.data-table td, .sheet-table td {
  border: 1px solid var(--border); padding: 7px 10px; vertical-align: top;
  line-height: 1.4; color: var(--txt);
}
.data-table tr:nth-child(even) td,
.sheet-table tr:nth-child(even) td { background: var(--row-alt); }
.data-table td[style*="background"],
.sheet-table td[style*="background"] { }

/* Scrollable large tables */
.table-scroll {
  max-height: 520px; overflow-y: auto;
  border: 1px solid var(--border);
}
.table-scroll .data-table,
.table-scroll .sheet-table { margin: 0; }
.table-scroll .data-table tr:first-child td,
.table-scroll .sheet-table tr:first-child td {
  position: sticky; top: 0; z-index: 2; box-shadow: 0 1px 0 var(--border2);
}

/* Wide sheets */
.wide-scroll { overflow-x: auto; -webkit-overflow-scrolling: touch; }
.wide-table { width: auto !important; white-space: nowrap; }
.wide-table td { white-space: nowrap; }
.table-wrap { overflow-x: auto; }
.sheet-table img { max-width:100%; height:auto; display:block; }

/* ── Print ───────────────────────────────────────────────────────────── */
/* ── Module tabs ─────────────────────────────────────────────────────── */
#module-tabs-wrap {
  border: 1px solid var(--border); box-shadow: var(--shadow); margin-bottom: 16px;
}
.mod-tab-bar {
  display: flex; flex-wrap: wrap; gap: 0;
  background: var(--surface2);
  border-bottom: 2px solid var(--border);
  overflow-x: auto;
}
.mod-tab {
  font-family: var(--cond); font-size: 12px; font-weight: 600;
  letter-spacing: .06em; text-transform: uppercase;
  padding: 10px 18px; border: none; background: transparent;
  color: var(--txt-muted); cursor: pointer;
  border-bottom: 2px solid transparent; margin-bottom: -2px;
  transition: color .1s, border-color .1s;
  white-space: nowrap;
}
.mod-tab:hover { color: var(--txt); }
.mod-tab.active { color: var(--red); border-bottom-color: var(--red); }
.mod-pane { display: none; padding: 0; background: var(--surface); }
.mod-pane.active { display: block; }

/* ── Sub-tabs (Overview / Changeover / BOM within a module pane) ─────── */
.sub-tab-bar {
  display: flex; gap: 0; background: var(--bg);
  border-bottom: 1px solid var(--border);
  padding: 0 16px;
}
.sub-tab {
  font-family: var(--cond); font-size: 11px; font-weight: 700;
  letter-spacing: .1em; text-transform: uppercase;
  padding: 8px 14px; border: none; background: transparent;
  color: var(--txt-muted); cursor: pointer;
  border-bottom: 2px solid transparent; margin-bottom: -1px;
  transition: color .1s, border-color .1s; white-space: nowrap;
}
.sub-tab:hover { color: var(--txt); }
.sub-tab.active { color: var(--red); border-bottom-color: var(--red); }
.sub-pane { display: none; padding: 16px; }
.sub-pane.active { display: block; }

/* ── Editable settings ───────────────────────────────────────────────── */
.sv-wrap { display: flex; align-items: center; gap: 5px; }
.sv-input {
  flex: 1; background: transparent; border: 1px solid transparent;
  font-family: var(--mono); font-size: 12px; color: var(--txt);
  padding: 3px 6px; border-radius: 3px; outline: none;
  transition: border-color .12s, background .12s;
}
.sv-input:focus { border-color: var(--red); background: rgba(204,0,0,.04); }
.sv-dot {
  width: 6px; height: 6px; border-radius: 50%;
  background: var(--red); flex-shrink: 0; opacity: 0;
  transition: opacity .15s;
}
.sv-dot.sv-modified { opacity: 1; }

/* ── PM icon column ─────────────────────────────────────────────────── */
.pm-icon { font-size: 18px; display: block; text-align: center; line-height: 1; }

/* ── HMI screen image ────────────────────────────────────────────────── */
.screen-img { background: #000; padding: 8px; border-radius: 4px; margin-bottom: 12px; }
.screen-img img { display: block; margin: 0 auto; max-height: 320px; width: auto; max-width: 100%; }

/* ── BOM table ───────────────────────────────────────────────────────── */
.bom-btn-cell { width: 36px; padding: 4px !important; text-align: center; }
.add-quote-btn {
  font-family: var(--mono); font-size: 11px; font-weight: 600;
  background: transparent; border: 1px solid var(--border);
  color: var(--txt-muted); padding: 2px 6px; cursor: pointer;
  border-radius: 3px; transition: all .1s;
}
.add-quote-btn:hover { border-color: var(--red); color: var(--red); }
.add-quote-btn.added { background: var(--red); border-color: var(--red); color: #fff; }

/* ── RFQ panel ───────────────────────────────────────────────────────── */
#rfq-fab {
  position: fixed; bottom: 24px; right: 24px; z-index: 500;
  background: var(--red); color: #fff; border: none;
  font-family: var(--cond); font-size: 13px; font-weight: 700;
  letter-spacing: .08em; text-transform: uppercase;
  padding: 10px 18px; cursor: pointer;
  box-shadow: 0 4px 16px rgba(204,0,0,.35);
  display: flex; align-items: center; gap: 8px;
}
#rfq-fab-count {
  background: #fff; color: var(--red);
  font-size: 11px; font-weight: 800; border-radius: 50%;
  width: 18px; height: 18px; display: inline-flex;
  align-items: center; justify-content: center;
}
#rfq-panel {
  position: fixed; bottom: 0; right: 0; z-index: 400;
  width: 380px; max-height: 60vh;
  background: var(--surface); border: 1px solid var(--border);
  border-top: 3px solid var(--red);
  box-shadow: -4px 0 24px rgba(0,0,0,.12);
  display: flex; flex-direction: column;
  transform: translateY(100%); transition: transform .2s ease;
}
#rfq-panel.rfq-open { transform: translateY(0); }
.rfq-panel-hdr {
  display: flex; align-items: center; padding: 12px 16px;
  border-bottom: 1px solid var(--border); flex-shrink: 0;
}
.rfq-panel-title {
  font-family: var(--cond); font-size: 13px; font-weight: 700;
  letter-spacing: .08em; text-transform: uppercase; flex: 1;
}
.rfq-send-btn {
  font-family: var(--cond); font-size: 11px; font-weight: 700;
  letter-spacing: .08em; text-transform: uppercase;
  background: var(--red); color: #fff; border: none;
  padding: 6px 14px; cursor: pointer; margin-right: 8px;
}
.rfq-send-btn:hover { opacity: .88; }
.rfq-close-btn {
  background: none; border: none; color: var(--txt-muted);
  font-size: 18px; cursor: pointer; line-height: 1; padding: 0 4px;
}
#rfq-list { flex: 1; overflow-y: auto; padding: 8px 0; }
.rfq-empty { padding: 20px 16px; font-family: var(--mono); font-size: 11px; color: var(--txt-dim); }
.rfq-item {
  display: flex; align-items: flex-start;
  padding: 8px 16px; border-bottom: 1px solid var(--border); gap: 8px;
}
.rfq-item-info { flex: 1; display: flex; flex-direction: column; gap: 2px; }
.rfq-pn  { font-family: var(--mono); font-size: 12px; color: var(--txt); font-weight: 500; }
.rfq-desc { font-size: 12px; color: var(--txt-muted); }
.rfq-cfg  { font-family: var(--mono); font-size: 10px; color: var(--txt-dim); }
.rfq-mod  { font-size: 10px; color: var(--red); letter-spacing: .04em; text-transform: uppercase; }
.rfq-remove { background: none; border: none; color: var(--txt-dim); cursor: pointer; font-size: 16px; padding: 0; }
.rfq-remove:hover { color: var(--red); }

/* ── HMI screens carousel ────────────────────────────────────────────── */
.hmi-carousel { margin: 16px 0; }
.hmi-carousel-bar {
  display: flex; align-items: center; gap: 12px;
  padding: 8px 12px; background: #1a1a1a; color: #fff;
  border: 1px solid rgba(204,0,0,.3); border-radius: 4px 4px 0 0;
  font-family: var(--cond);
}
.hmi-prev, .hmi-next {
  background: var(--red); color: #fff; border: none;
  width: 32px; height: 32px; border-radius: 4px;
  font-size: 22px; line-height: 1; cursor: pointer;
  transition: background .12s, transform .08s;
}
.hmi-prev:hover, .hmi-next:hover { background: #e60000; }
.hmi-prev:active, .hmi-next:active { transform: scale(.94); }
.hmi-counter {
  margin-left: auto; font-size: 13px; letter-spacing: .04em;
  color: rgba(255,255,255,.75);
}
.hmi-counter .hmi-cur { color: #fff; font-weight: 700; }
.hmi-thumbs {
  display: flex; flex-wrap: wrap; gap: 4px;
  padding: 6px 8px; background: #f4f4f4;
  border: 1px solid rgba(204,0,0,.25); border-top: none;
}
.hmi-thumb {
  background: #fff; color: #555; border: 1px solid rgba(0,0,0,.15);
  font-family: var(--cond); font-size: 11px; font-weight: 600;
  text-transform: uppercase; letter-spacing: .04em;
  padding: 5px 10px; cursor: pointer; border-radius: 3px;
  transition: background .1s, color .1s, border-color .1s;
}
.hmi-thumb:hover { background: rgba(204,0,0,.08); color: var(--red); }
.hmi-thumb.active {
  background: var(--red); color: #fff; border-color: var(--red);
}
.hmi-slides {
  border: 1px solid rgba(204,0,0,.25); border-top: none;
  border-radius: 0 0 4px 4px;
  padding: 16px; background: #fff;
}
.hmi-slide { display: none; }
.hmi-slide.active { display: block; }
.hmi-slide-title {
  margin: 0 0 12px; font-family: var(--cond); font-size: 15px;
  font-weight: 700; color: var(--red);
  text-transform: uppercase; letter-spacing: .06em;
}
.hmi-slide .screen-img { margin-bottom: 12px; text-align: center; }
.hmi-slide .screen-img img { max-width: 100%; height: auto; display: inline-block; }

/* ── Print ───────────────────────────────────────────────────────────── */
@media print {
  #sidebar, #topbar, #rfq-fab, #rfq-panel { display: none; }
  #main { padding: 0; }
  .grp-section, .sheet-panel { break-inside: avoid; box-shadow: none; }
  .grp-body, .sheet-panel-body { display: block !important; }
  .table-scroll { max-height: none; overflow: visible; }
  .grp-section-hdr, .sheet-panel-hdr { print-color-adjust: exact; -webkit-print-color-adjust: exact; }
  .mod-pane { display: block !important; }
  .sub-pane { display: block !important; }
  .mod-tab-bar, .sub-tab-bar { display: none; }
  .hmi-carousel-bar, .hmi-thumbs { display: none; }
  .hmi-slide { display: block !important; page-break-inside: avoid; margin-bottom: 16px; }
  .hmi-slides { border: none; padding: 0; }
}
"""

JS = """
(function() {

  /* ── Group accordion ──────────────────────────────────────────────── */
  document.querySelectorAll('.grp-section-hdr').forEach(function(hdr) {
    hdr.addEventListener('click', function() {
      var grp = hdr.parentElement;
      grp.classList.toggle('collapsed');
      // Sync sidebar group nav
      var gid = grp.dataset.grpId;
      if (gid) {
        var navGrp = document.querySelector('.grp-nav[data-grp="' + gid + '"]');
        if (navGrp) navGrp.classList.toggle('nav-closed', grp.classList.contains('collapsed'));
      }
    });
  });

  /* ── Sheet panel accordion ────────────────────────────────────────── */
  document.querySelectorAll('.sheet-panel-hdr').forEach(function(hdr) {
    hdr.addEventListener('click', function() {
      hdr.parentElement.classList.toggle('collapsed');
    });
  });

  /* ── Sidebar group nav toggle ─────────────────────────────────────── */
  document.querySelectorAll('.grp-nav-hdr').forEach(function(hdr) {
    hdr.addEventListener('click', function() {
      var nav = hdr.parentElement;
      nav.classList.toggle('nav-closed');
      // Sync main content group
      var gid = nav.dataset.grp;
      if (gid) {
        var grp = document.querySelector('.grp-section[data-grp-id="' + gid + '"]');
        if (grp) grp.classList.toggle('collapsed', nav.classList.contains('nav-closed'));
      }
    });
  });

  /* ── Sidebar links: expand group + panel, scroll ─────────────────── */
  function expandToTarget(target) {
    // Expand containing group
    var grp = target.closest('.grp-section');
    if (grp && grp.classList.contains('collapsed')) {
      grp.classList.remove('collapsed');
      var gid = grp.dataset.grpId;
      if (gid) {
        var navGrp = document.querySelector('.grp-nav[data-grp="' + gid + '"]');
        if (navGrp) navGrp.classList.remove('nav-closed');
      }
    }
    // Expand containing sheet panel
    var panel = target.closest('.sheet-panel');
    if (panel && panel.classList.contains('collapsed')) panel.classList.remove('collapsed');
  }

  /* ── Sidebar link handler ─────────────────────────────────────────── */
  document.querySelectorAll('#sidebar a[href^="#"]').forEach(function(a) {
    var tid = a.getAttribute('href').slice(1);
    a.addEventListener('click', function(e) {
      e.preventDefault();
      var target = document.getElementById(tid);
      if (!target) return;
      // Module slide: expand parent panels then activate carousel position
      if (target.classList.contains('hmi-slide')) {
        expandToTarget(target);
        var car = target.closest('.hmi-carousel');
        var idx = parseInt(target.dataset.idx, 10);
        if (car) {
          var thumb = car.querySelector('.hmi-thumb[data-idx="' + idx + '"]');
          if (thumb) thumb.click();
          car.scrollIntoView({ behavior: 'smooth', block: 'start' });
        }
      } else {
        expandToTarget(target);
        target.scrollIntoView({ behavior: 'smooth', block: 'start' });
      }
      document.querySelectorAll('#sidebar a').forEach(function(x){ x.classList.remove('active'); });
      a.classList.add('active');
    });
  });

  /* ── Editable settings ────────────────────────────────────────────── */
  document.querySelectorAll('.sv-input').forEach(function(inp) {
    inp.addEventListener('input', function() {
      var dot = inp.nextElementSibling;
      if (dot) dot.classList.toggle('sv-modified', inp.value !== inp.dataset.orig);
    });
  });
  // Reset on double-click
  document.querySelectorAll('.sv-input').forEach(function(inp) {
    inp.addEventListener('dblclick', function() {
      inp.value = inp.dataset.orig;
      var dot = inp.nextElementSibling;
      if (dot) dot.classList.remove('sv-modified');
    });
  });

  /* ── RFQ / Add to Quote ───────────────────────────────────────────── */
  var rfqItems = [];
  function rfqCount() { return rfqItems.length; }
  function updateRFQBadge() {
    var b = document.getElementById('rfq-fab-count');
    if (b) b.textContent = rfqCount();
    var panel = document.getElementById('rfq-panel');
    if (panel) renderRFQPanel();
  }
  function renderRFQPanel() {
    var list = document.getElementById('rfq-list');
    if (!list) return;
    if (rfqItems.length === 0) {
      list.innerHTML = '<div class="rfq-empty">No items added yet.</div>';
      return;
    }
    list.innerHTML = rfqItems.map(function(item, i) {
      return '<div class="rfq-item">' +
        '<div class="rfq-item-info">' +
          '<span class="rfq-pn">' + item.pn + '</span>' +
          '<span class="rfq-desc">' + item.desc + '</span>' +
          (item.cfg ? '<span class="rfq-cfg">' + item.cfg + '</span>' : '') +
          '<span class="rfq-mod">' + item.module + '</span>' +
        '</div>' +
        '<button class="rfq-remove" onclick="removeRFQItem(' + i + ')">×</button>' +
      '</div>';
    }).join('');
  }
  window.removeRFQItem = function(i) {
    rfqItems.splice(i, 1);
    updateRFQBadge();
  };
  window.toggleRFQ = function() {
    var panel = document.getElementById('rfq-panel');
    if (panel) { panel.classList.toggle('rfq-open'); renderRFQPanel(); }
  };
  window.sendRFQ = function() {
    var topbar   = document.getElementById('topbar');
    var serial   = (topbar ? (topbar.querySelector('.tb-sn') || {}).textContent || '' : '').trim();
    var customer = (topbar ? (topbar.querySelector('.tb-customer') || {}).textContent || '' : '').trim();
    var today    = new Date().toISOString().slice(0,10);
    var subject  = encodeURIComponent(serial + ' - ' + customer + ' - RFQ - ' + today);
    var lines    = rfqItems.map(function(it) {
      return 'Part #: '      + it.pn + '\\n' +
             'Description: ' + (it.desc || '') + '\\n' +
             'Config: '      + (it.cfg  || 'Standard') + '\\n' +
             'Qty: '         + (it.qty  || '1') + '\\n' +
             'Module: '      + it.module + '\\n';
    });
    var body = encodeURIComponent(
      'Hello,\\n\\nPlease provide a quote for the following parts:\\n\\n' +
      lines.join('\\n') +
      '\\nRegards'
    );
    window.location.href = 'mailto:parts@polypack.com?subject=' + subject + '&body=' + body;
  };
  document.querySelectorAll('.add-quote-btn').forEach(function(btn) {
    btn.addEventListener('click', function() {
      var item = {
        pn:     btn.dataset.pn     || '',
        desc:   btn.dataset.desc   || '',
        cfg:    btn.dataset.cfg    || '',
        qty:    btn.dataset.qty    || '1',
        module: btn.dataset.module || ''
      };
      // Avoid duplicates
      if (!rfqItems.some(function(x){ return x.pn === item.pn; })) {
        rfqItems.push(item);
        btn.textContent = '✓';
        btn.classList.add('added');
        updateRFQBadge();
      }
    });
  });

  // ── HMI screens carousel ────────────────────────────────────────────
  document.querySelectorAll('.hmi-carousel').forEach(function(car) {
    var slides = car.querySelectorAll('.hmi-slide');
    var thumbs = car.querySelectorAll('.hmi-thumb');
    var prev   = car.querySelector('.hmi-prev');
    var next   = car.querySelector('.hmi-next');
    var curEl  = car.querySelector('.hmi-cur');
    var total  = slides.length;
    if (!total) return;
    function show(i) {
      i = (i + total) % total;
      slides.forEach(function(s, j) { s.classList.toggle('active', j === i); });
      thumbs.forEach(function(t, j) { t.classList.toggle('active', j === i); });
      car.dataset.current = i;
      if (curEl) curEl.textContent = (i + 1);
      // Scroll the active thumb into view (horizontally)
      var at = thumbs[i];
      if (at && at.scrollIntoView) at.scrollIntoView({block: 'nearest', inline: 'nearest'});
    }
    if (prev) prev.addEventListener('click', function() { show(parseInt(car.dataset.current, 10) - 1); });
    if (next) next.addEventListener('click', function() { show(parseInt(car.dataset.current, 10) + 1); });
    thumbs.forEach(function(t) {
      t.addEventListener('click', function() { show(parseInt(t.dataset.idx, 10)); });
    });
    // Keyboard arrows when the carousel (or anything inside it) has focus
    car.tabIndex = 0;
    car.addEventListener('keydown', function(e) {
      if (e.key === 'ArrowLeft')  { show(parseInt(car.dataset.current, 10) - 1); e.preventDefault(); }
      if (e.key === 'ArrowRight') { show(parseInt(car.dataset.current, 10) + 1); e.preventDefault(); }
    });
  });
})();
"""


def sheet_id(name: str) -> str:
    return "s-" + re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")


def segments_to_html(segments: list[dict]) -> str:
    """Render flat segments (non-module sheets). subsection types are rendered inline."""
    parts = []
    for seg in segments:
        if seg["type"] == "html":
            parts.append(seg["content"])
        elif seg["type"] == "subsection":
            # Module sheet rendered without sub-tabs (fallback / settings context)
            parts.append(f'<div class="sub-section"><h3>{seg["label"]}</h3>{seg["html"]}</div>')
        elif seg["type"] == "img":
            if seg.get("file"):
                parts.append(f'<img src="assets/{url_quote(seg["file"])}" alt="{seg.get("alt","")}">')
            else:
                parts.append(f'<p><em>(image export failed: {seg.get("alt","")})</em></p>')
    return "\n".join(parts)


def module_pane_html(segments: list[dict], module_name: str) -> str:
    """
    Render a module's segments stacked: plain HTML first, then each subsection
    with its label as an h3 heading.  No carousel here — sections within a
    single module are shown together.  The outer carousel (between modules)
    is built in build_html.
    """
    return segments_to_html(segments)


def build_html(sheet_segments: list[dict], machine_info: dict) -> str:
    """
    sheet_segments: [{name, label, type, group, segments}, ...]
    type:  'module'   → tabbed pane within modules panel
           'settings' → editable inputs
           ''         → regular content
    group: GROUP_ORDER key determining which accordion section the sheet lives in
    """
    serial   = machine_info.get("serial", "")
    mtype    = machine_info.get("type", "")
    customer = machine_info.get("customer", "")

    # Bucket sheets into groups, preserving order within each group
    groups: dict[str, list] = {}
    for s in sheet_segments:
        grp = s.get("group") or "OVERVIEW"
        groups.setdefault(grp, []).append(s)

    # ── Sidebar nav (one .grp-nav block per group) ────────────────────────────
    nav_parts = []
    for grp_key, grp_label in GROUP_ORDER:
        sheets = groups.get(grp_key, [])
        if not sheets:
            continue
        mods    = [s for s in sheets if s.get("type") == "module"]
        non_mod = [s for s in sheets if s.get("type") != "module"]
        nav_parts.append(
            f'<div class="grp-nav" data-grp="{grp_key}">'
            f'<span class="grp-nav-hdr">{grp_label}</span>'
            f'<div class="grp-nav-sub">'
        )
        for s in non_mod:
            nav_parts.append(f'<a href="#{sheet_id(s["name"])}">{s["label"]}</a>')
        if mods:
            nav_parts.append('<span class="mod-nav-label">Modules</span>')
            nav_parts.append('<div class="mod-nav-list">')
            for s in mods:
                nav_parts.append(f'<a href="#{sheet_id(s["name"])}">{s["label"]}</a>')
            nav_parts.append('</div>')
        nav_parts.append('</div></div>')

    # ── Main content (one .grp-section accordion per group) ───────────────────
    content_parts = []
    for grp_key, grp_label in GROUP_ORDER:
        sheets = groups.get(grp_key, [])
        if not sheets:
            continue
        mods    = [s for s in sheets if s.get("type") == "module"]
        non_mod = [s for s in sheets if s.get("type") != "module"]

        panels = []

        # Non-module sheets — each gets its own collapsible panel
        for s in non_mod:
            sid   = sheet_id(s["name"])
            has_subs = any(seg["type"] == "subsection" for seg in s["segments"])
            body  = module_pane_html(s["segments"], s["name"]) if has_subs \
                    else segments_to_html(s["segments"])
            stype = s.get("type", "")
            badge = ('<span class="sheet-panel-hdr-badge">editable</span>'
                     if stype == "settings" else "")
            panels.append(
                f'<div class="sheet-panel collapsed" id="{sid}">'
                f'<div class="sheet-panel-hdr">'
                f'<span class="sheet-panel-hdr-title">{s["label"]}</span>{badge}'
                f'</div>'
                f'<div class="sheet-panel-body">{body}</div>'
                f'</div>'
            )

        # Modules — carousel within a single collapsible panel.
        # Each module is one slide; subsections inside a module render stacked.
        if mods:
            total = len(mods)
            thumbs = "".join(
                f'<button class="hmi-thumb{" active" if i == 0 else ""}" '
                f'type="button" data-idx="{i}">{s["label"]}</button>'
                for i, s in enumerate(mods)
            )
            slides = "".join(
                f'<div class="hmi-slide{" active" if i == 0 else ""}" '
                f'data-idx="{i}" id="{sheet_id(s["name"])}">'
                f'<h3 class="hmi-slide-title">{s["label"]}</h3>'
                f'{module_pane_html(s["segments"], s["name"])}'
                f'</div>'
                for i, s in enumerate(mods)
            )
            mod_wrap = (
                '<div class="hmi-carousel module-carousel" data-current="0" id="module-carousel">'
                '<div class="hmi-carousel-bar">'
                '<button class="hmi-prev" type="button" aria-label="Previous module">&lsaquo;</button>'
                f'<span class="hmi-counter"><span class="hmi-cur">1</span> / '
                f'<span class="hmi-total">{total}</span></span>'
                '<button class="hmi-next" type="button" aria-label="Next module">&rsaquo;</button>'
                '</div>'
                f'<div class="hmi-thumbs">{thumbs}</div>'
                f'<div class="hmi-slides">{slides}</div>'
                '</div>'
            )
            panels.append(
                '<div class="sheet-panel collapsed module-panel" id="grp-modules">'
                '<div class="sheet-panel-hdr">'
                '<span class="sheet-panel-hdr-title">Modules</span>'
                '</div>'
                '<div class="sheet-panel-body">' + mod_wrap + '</div>'
                '</div>'
            )

        content_parts.append(
            f'<div class="grp-section collapsed" data-grp-id="{grp_key}">'
            f'<div class="grp-section-hdr">'
            f'<span class="grp-section-hdr-title">{grp_label}</span>'
            f'</div>'
            f'<div class="grp-body">'
            + "".join(panels) +
            f'</div>'
            f'</div>'
        )

    content_html = "".join(content_parts)
    nav_html     = "".join(nav_parts)
    title = "Manual \u2014 " + mtype + (" \u2014 S/N " + serial if serial else "")

    # Machine info grid
    info_grid = (
        '<div class="machine-info-grid">'
        '<div class="mi-cell"><div class="mi-label">Customer</div>'
        '<div class="mi-value">' + (customer or "&mdash;") + '</div></div>'
        '<div class="mi-cell"><div class="mi-label">Machine Type</div>'
        '<div class="mi-value">' + (mtype or "&mdash;") + '</div></div>'
        '<div class="mi-cell"><div class="mi-label">Serial Number</div>'
        '<div class="mi-value accent small">' + (serial or "&mdash;") + '</div></div>'
        '</div>'
    )

    return (
        "<!DOCTYPE html>\n<html lang='en'>\n<head>\n"
        "  <meta charset='UTF-8'>\n"
        "  <meta name='viewport' content='width=device-width,initial-scale=1'>\n"
        "  <title>" + title + "</title>\n"
        "  <style>" + CSS + "</style>\n"
        "</head>\n<body>\n"
        "<div id='topbar'>\n"
        "  <div class='tb-brand'>Poly<span>pack</span></div>\n"
        "  <div class='tb-divider'></div>\n"
        "  <div class='tb-machine'>" + mtype + "</div>\n"
        "  <span class='tb-sn'>" + serial + "</span>\n"
        "  <div class='tb-customer'>" + customer + "</div>\n"
        "</div>\n"
        "<div id='app'>\n"
        "<aside id='sidebar'>\n"
        "  <span class='sb-section-label'>Navigation</span>\n"
        "  <nav>" + nav_html + "</nav>\n"
        "</aside>\n"
        "<div id='main'>\n"
        + info_grid
        + content_html +
        "\n</div>\n</div>\n"
        "<button id='rfq-fab' onclick='toggleRFQ()'>"
        "Parts Quote <span id='rfq-fab-count'>0</span></button>\n"
        "<div id='rfq-panel'>\n"
        "  <div class='rfq-panel-hdr'>\n"
        "    <span class='rfq-panel-title'>Parts Quote</span>\n"
        "    <button class='rfq-send-btn' onclick='sendRFQ()'>Send RFQ</button>\n"
        "    <button class='rfq-close-btn' onclick='toggleRFQ()'>&#215;</button>\n"
        "  </div>\n"
        "  <div id='rfq-list'></div>\n"
        "</div>\n"
        "<script>" + JS + "</script>\n"
        "</body>\n</html>\n"
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# HTM mode — post-process Excel "Save As HTML" export
# ---------------------------------------------------------------------------

try:
    from bs4 import BeautifulSoup, Comment, Tag as BsTag
    BS4_OK = True
except ImportError:
    BS4_OK = False

# Sheet names in display order for --htm mode
HTM_ALWAYS = [
    "SUMMARY", "FTD",
    # modules inserted after FTD (see MODULE_SHEETS_ORDERED below)
    "ELE", "FAULTS", "PM", "TS", "HMI", "SCREENS",
    "PHYSICAL", "LUBRICATION",
    # VFD appended dynamically from var_PLC
]
MODULE_SHEETS_ORDERED = [
    "INFEED CONVEYOR", "STAGING CLAMP", "FILM CRADLE", "FILM DELIVERY",
    "TRANSFER CONVEYOR", "WRAP ARM", "TUNNEL CONVEYOR", "HEAT TUNNEL",
]
# Sheets that need horizontal scrolling (many columns)
WIDE_SHEETS = {"SCREENS", "ELE", "tbl_BUTTONS"}

# Normalize var_PLC value → sheet name
VFD_MAP = {
    "SCHNEIDER": "SCHNEIDER",
    "AB": "AB", "ALLEN-BRADLEY": "AB", "ALLEN BRADLEY": "AB",
    "POWERFLEX": "POWERFLEX", "POWERFLEX 525": "POWERFLEX",
}

# Row text patterns that indicate a sheet-header block to remove from web output
HEADER_KEYWORDS = {
    "TECHNICAL MANUAL", "ENGINEERING MACHINE SUMMARY",
    "SERIAL NUMBER", "SERIAL NUMBER:", "TYPE", "TYPE:",
    "CUSTOMER", "CUSTOMER:", "DELIVERY DATE", "DELIVERY DATE:",
}

MSO_PROP_RE = re.compile(r'\bmso-[^;:]+:[^;]+;?\s*')


def _strip_mso(style: str) -> str:
    cleaned = MSO_PROP_RE.sub("", style)
    return re.sub(r";+", ";", cleaned).strip("; ")


def parse_frameset_htm(htm_path: Path) -> dict[str, str]:
    """Return {SHEET_NAME: 'sheetXXX.htm'} from main frameset."""
    text = htm_path.read_text(encoding="windows-1252", errors="replace")
    names = re.findall(r'c_rgszSh\[\d+\]\s*=\s*"([^"]+)"', text)
    links = re.findall(r'href="[^"]*/(sheet\d+\.htm)"', text)
    if len(names) != len(links):
        raise ValueError(f"Sheet name/link mismatch: {len(names)} vs {len(links)}")
    return {n.replace("\xa0", " ").strip(): lnk for n, lnk in zip(names, links)}


def extract_machine_info_htm(summary_path: Path) -> dict:
    """Read serial, type, customer, and var_PLC from SUMMARY sheet HTML."""
    text = summary_path.read_text(encoding="windows-1252", errors="replace")
    info = {"serial": "", "type": "", "customer": "", "plc": ""}
    for field, anchor in [("serial", "SerialNo"), ("type", "MachType"), ("customer", "Customer")]:
        m = re.search(rf'<a\s+name={anchor}>([^<]+)</a>', text, re.IGNORECASE)
        if m:
            info[field] = m.group(1).strip()
    # var_PLC: find a cell whose text is "var_PLC" and grab the next TD's value
    soup = BeautifulSoup(text, "html.parser")
    for td in soup.find_all("td"):
        val = td.get_text(strip=True)
        if val.lower() == "var_plc":
            nxt = td.find_next_sibling("td")
            if nxt:
                info["plc"] = nxt.get_text(strip=True)
            break
    return info


def build_sheet_list_htm(sheet_map: dict, machine_info: dict) -> list[tuple[str, str]]:
    """Return ordered [(sheet_name, label)] for the web manual."""
    plc_raw = machine_info.get("plc", "").upper().replace(" ", " ")
    vfd_sheet = VFD_MAP.get(plc_raw, "")

    ordered = []
    for name in HTM_ALWAYS:
        if name in sheet_map:
            ordered.append(name)
        if name == "FTD":
            for mod in MODULE_SHEETS_ORDERED:
                if mod in sheet_map:
                    ordered.append(mod)

    if vfd_sheet and vfd_sheet in sheet_map:
        ordered.append(vfd_sheet)
    elif not vfd_sheet:
        # Fallback: include whichever VFD sheet is present
        for v in ["SCHNEIDER", "AB", "POWERFLEX"]:
            if v in sheet_map:
                ordered.append(v)
                break

    return [(n, LABELS.get(n, n.title())) for n in ordered]


def clean_sheet_htm(sheet_path: Path, sheet_name: str, assets_dir: Path) -> str:
    """
    Parse one sheet HTML file and return cleaned table HTML.
    - Removes hidden rows, IE conditionals, VML, MSO boilerplate
    - Removes sheet-header rows (TECHNICAL MANUAL / SERIAL NUMBER etc.)
    - Fixes image src to point at assets/
    - Removes fixed TD width/height constraints on image cells
    - Strips <col> fixed widths
    """
    text = sheet_path.read_text(encoding="windows-1252", errors="replace")
    # Strip IE conditional blocks
    text = re.sub(r'<!\[if[^\]]*\]>.*?<!\[endif\]>', '', text, flags=re.DOTALL)
    text = re.sub(r'<!--.*?-->', '', text, flags=re.DOTALL)

    soup = BeautifulSoup(text, "html.parser")

    for tag in soup.find_all(["script", "style", "link", "meta"]):
        tag.decompose()
    for c in soup.find_all(string=lambda s: isinstance(s, Comment)):
        c.extract()

    # Remove hidden rows
    for tr in soup.find_all("tr"):
        s = tr.get("style", "").replace(" ", "")
        if "display:none" in s or tr.get("height") == "0":
            tr.decompose()

    table = soup.find("table")
    if not table:
        return ""

    # Remove <col> fixed widths (makes table responsive)
    for col in table.find_all("col"):
        col.decompose()

    # Remove sheet-header rows (TECHNICAL MANUAL / SERIAL NUMBER / etc.)
    for tr in table.find_all("tr"):
        cells = [td.get_text(" ", strip=True).upper().rstrip(":") for td in tr.find_all(["td", "th"])]
        first = cells[0] if cells else ""
        if first in HEADER_KEYWORDS:
            tr.decompose()

    # Fix images and relax containing TD constraints
    for img in table.find_all("img"):
        src = img.get("src", "")
        # Robust filename extraction — handles any path separator or prefix
        fname = src.replace("\\", "/").split("/")[-1] if src else ""
        if fname:
            img["src"] = "assets/" + fname
        img["style"] = "max-width:100%;height:auto;display:block;"
        img.attrs.pop("width", None)
        img.attrs.pop("height", None)
        # Remove fixed width/height from the parent TD
        parent = img.find_parent("td")
        if parent:
            parent.attrs.pop("width", None)
            parent.attrs.pop("height", None)
            st = parent.get("style", "")
            st = re.sub(r'\bwidth\s*:[^;]+;?', '', st)
            st = re.sub(r'\bheight\s*:[^;]+;?', '', st)
            st = st.strip("; ")
            if st:
                parent["style"] = st
            else:
                parent.attrs.pop("style", None)

    # Strip MSO CSS from all tags, remove class/VML attrs
    for tag in table.find_all(True):
        if tag.get("style"):
            tag["style"] = _strip_mso(tag["style"]) or None
            if not tag["style"]:
                del tag["style"]
        for attr in ("class", "x:str", "x:num", "x:fmla", "v:shapes"):
            tag.attrs.pop(attr, None)

    is_wide = sheet_name.upper() in WIDE_SHEETS
    table["class"] = "sheet-table wide-table" if is_wide else "sheet-table"

    wrapper = "wide-scroll" if is_wide else "table-wrap"
    return f'<div class="{wrapper}">{table}</div>'


def run_htm_mode(htm_path: Path, out_dir: Path) -> None:
    """Main logic for --htm mode."""
    if not BS4_OK:
        sys.exit("beautifulsoup4 not found.  Run:  python -m pip install beautifulsoup4")

    # Locate _files directory from the frameset's File-List link
    raw = htm_path.read_text(encoding="windows-1252", errors="replace")
    fl  = re.search(r'href="([^"]+)/filelist\.xml"', raw, re.IGNORECASE)
    files_dir = htm_path.parent / (unquote(fl.group(1)) if fl else htm_path.stem + "_files")
    if not files_dir.exists():
        sys.exit(f"Files directory not found: {files_dir}")

    out_dir.mkdir(parents=True, exist_ok=True)
    assets_dir = out_dir / "assets"
    assets_dir.mkdir(exist_ok=True)

    print(f"Reading frameset: {htm_path.name}")
    sheet_map = parse_frameset_htm(htm_path)
    print(f"  {len(sheet_map)} sheets found")

    # Machine info from SUMMARY
    machine_info = {}
    if "SUMMARY" in sheet_map:
        machine_info = extract_machine_info_htm(files_dir / sheet_map["SUMMARY"])
    print(f"  Machine: {machine_info}")
    print(f"  PLC: {machine_info.get('plc') or '(not set — will use first available VFD sheet)'}")

    # Diagnostic: show what's in the _files directory
    all_files = list(files_dir.iterdir())
    print(f"  Files dir: {files_dir}")
    print(f"  Total files in _files dir: {len(all_files)}")
    if all_files:
        sample = [f.name for f in all_files[:10]]
        print(f"  Sample filenames: {sample}")

    # Copy all images to assets/ (PNG and JPEG, case-insensitive on Windows)
    count = 0
    for f in files_dir.iterdir():
        if f.suffix.lower() in (".png", ".jpg", ".jpeg", ".gif") and f.is_file():
            shutil.copy2(f, assets_dir / f.name)
            count += 1
    print(f"  Copied {count} images to assets/")

    # Build ordered sheet list
    sheet_list = build_sheet_list_htm(sheet_map, machine_info)
    print(f"  Sheets to render ({len(sheet_list)}): {', '.join(n for n,_ in sheet_list)}\n")

    # Parse each sheet
    sheet_segments = []
    for name, label in sheet_list:
        sheet_file = files_dir / sheet_map[name]
        if not sheet_file.exists():
            print(f"  Skip (missing): {name}")
            continue
        print(f"  Parsing: {name}")
        html = clean_sheet_htm(sheet_file, name, assets_dir)
        if html:
            # Single-segment sheets: wrap in one html segment for build_html
            sheet_segments.append({
                "name": name, "label": label,
                "segments": [{"type": "html", "content": html}],
            })

    print("\nBuilding manual.html …")
    out_file = out_dir / "manual.html"
    out_file.write_text(build_html(sheet_segments, machine_info), encoding="utf-8")
    print(f"Done → {out_file}  ({out_file.stat().st_size // 1024} KB)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Generate HTML manual from Excel workbook")
    grp = ap.add_mutually_exclusive_group(required=True)
    grp.add_argument("--htm",  help="Path to Excel 'Save As HTML' frameset .htm file")
    grp.add_argument("--xlsx", help="Path to Excel workbook (requires img_* named ranges)")
    ap.add_argument("--out", required=True, help="Output directory")
    args = ap.parse_args()

    out_dir = Path(args.out).resolve()

    if args.htm:
        htm_path = Path(args.htm).resolve()
        if not htm_path.exists():
            sys.exit(f"File not found: {htm_path}")
        run_htm_mode(htm_path, out_dir)

    else:
        xlsx_path = Path(args.xlsx).resolve()
        if not xlsx_path.exists():
            sys.exit(f"File not found: {xlsx_path}")

        out_dir.mkdir(parents=True, exist_ok=True)
        # assets/ is created by VBA next to the xlsx; if --out is the xlsx dir they're the same
        assets_dir = out_dir / "assets"
        assets_dir.mkdir(exist_ok=True)

        # ── Debug log (persists after terminal closes) ──────────────────────
        log_path = out_dir / "polish_debug.log"
        log_lines = [f"polish.py debug log — {datetime.datetime.now()}",
                     f"xlsx : {xlsx_path}",
                     f"out  : {out_dir}",
                     f"assets dir exists: {assets_dir.exists()}"]
        all_pngs = sorted(
            f.name for f in assets_dir.iterdir() if f.suffix.lower() == ".png"
        ) if assets_dir.exists() else []
        log_lines.append(f"\nAll PNGs in assets/ ({len(all_pngs)}):")
        for p in all_pngs:
            log_lines.append(f"  {p}")

        def _log(msg):
            print(msg)
            log_lines.append(msg)

        print(f"Opening (openpyxl): {xlsx_path.name}")
        wb_opxl = openpyxl.load_workbook(xlsx_path, data_only=True)

        machine_info = get_machine_info_opxl(wb_opxl)
        _log(f"\nMachine: {machine_info}")

        sheet_items = read_print_table_opxl(wb_opxl)

        # Drop VFD sheets that don't match the active PLC architecture.
        active_vfd = (machine_info.get("plc") or "").upper()
        if active_vfd in VFD_SHEETS:
            before = len(sheet_items)
            sheet_items = [
                it for it in sheet_items
                if it["name"].upper() not in VFD_SHEETS
                or it["name"].upper() == active_vfd
            ]
            dropped = before - len(sheet_items)
            if dropped:
                _log(f"VFD filter: keeping {active_vfd}, dropped {dropped} other VFD sheet(s)")

        _log(f"Sheets ({len(sheet_items)}): {', '.join(i['name'] for i in sheet_items)}\n")

        sheet_segments = []
        for item in sheet_items:
            name  = item["name"]
            stype = item["type"]
            label = LABELS.get(name.upper(), name.title())
            _log(f"  Processing: {name} [{stype or 'regular'}]")
            segs = process_sheet_opxl(name, wb_opxl, assets_dir, sheet_type=stype)

            # Determine navigation group: A1 note GROUP= > default map > OVERVIEW
            grp = ""
            try:
                ws_meta = wb_opxl[name] if name in wb_opxl.sheetnames else None
                if ws_meta:
                    m = _read_sheet_meta(ws_meta)
                    grp = m.get("GROUP", "")
            except Exception:
                pass
            if not grp:
                grp = _DEFAULT_GROUP.get(name.upper(), "")
            if not grp:
                grp = "OVERVIEW" if stype == "module" else "OVERVIEW"

            sheet_segments.append({"name": name, "label": label,
                                   "type": stype, "group": grp, "segments": segs})
            _log(f"    {len(segs)} segment(s) [group={grp}]")

        _log("\nBuilding manual.html …")
        # Write debug log before building HTML in case of crash
        log_path.write_text("\n".join(log_lines), encoding="utf-8")
        out_file = out_dir / "manual.html"
        out_file.write_text(build_html(sheet_segments, machine_info), encoding="utf-8")
        print(f"Done → {out_file}  ({out_file.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()